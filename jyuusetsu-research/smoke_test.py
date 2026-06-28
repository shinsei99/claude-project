"""UI を起動せずパイプライン各段を検証するスモークテスト。"""
import os

from models.property_data import create_property_data, merge
from services import (
    comment_service,
    excel_export_service,
    pdf_export_service,
    registry_service,
)
from utils import parser

BASE = os.path.dirname(os.path.abspath(__file__))

# 1) PropertyData
data = create_property_data()
assert "所在地" in data and data["所在地"] == ""

# 2) マージ（空/不正キーを無視）
merge(data, {"所在地": " 東京都千代田区丸の内1-1-1 ", "用途地域": "商業地域",
             "建ぺい率": "80%", "容積率": "800%", "不明キー": "x", "地番": ""})
assert data["所在地"] == "東京都千代田区丸の内1-1-1"
assert "不明キー" not in data
assert data["地番"] == ""  # 空文字は上書きしない

# 3) 登記簿パーサ（テキスト直接）
land_text = "所　在  千代田区丸の内一丁目\n地　番  1番1\n地　目  宅地\n地　積  123.45㎡\n所有者  山田太郎\n抵当権設定"
land = parser.parse_land(land_text)
assert land["地目"] == "宅地", land
assert land["地積"].startswith("123"), land
assert "抵当権" in land["抵当権"], land
print("land parse:", land)

# 4) registry_service 統合
reg = registry_service.parse_registry(None, None)
assert isinstance(reg, dict)

# 5) コメント生成
merge(data, {"最寄駅": "東京駅（JR山手線）", "駅距離": "約 350m"})
comment = comment_service.generate_comment(data)
assert "商業地域" in comment and len(comment) > 50
print("comment:", comment)

# 6) Excel 出力（テンプレ自動生成含む）
tpl = os.path.join(BASE, "templates", "jyuusetsu_template.xlsx")
out_xlsx = os.path.join(BASE, "reports", "jyuusetsu_draft.xlsx")
excel_export_service.export_excel(data, comment, tpl, out_xlsx)
assert os.path.exists(tpl) and os.path.exists(out_xlsx)
from openpyxl import load_workbook
ws = load_workbook(out_xlsx).active
assert ws["B2"].value == "東京都千代田区丸の内1-1-1", ws["B2"].value
assert ws["B10"].value == "商業地域", ws["B10"].value
print("excel B2/B10/B11:", ws["B2"].value, ws["B10"].value, ws["B11"].value)

# 7) PDF 出力
out_pdf = os.path.join(BASE, "reports", "jyuusetsu_draft.pdf")
pdf_export_service.export_pdf(data, comment, out_pdf)
assert os.path.exists(out_pdf) and os.path.getsize(out_pdf) > 1000
print("pdf bytes:", os.path.getsize(out_pdf))

print("\nALL SMOKE TESTS PASSED")
