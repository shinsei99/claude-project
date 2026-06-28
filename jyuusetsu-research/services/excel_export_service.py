"""重説テンプレート(Excel)へのセルマッピング書き込み。

- テンプレートが無ければ自動生成する（A 列ラベル / B 列に値）。
- CELL_MAP に従い PropertyData を書き込み jyuusetsu_draft.xlsx を出力する。

セルマッピングは下記 CELL_MAP を編集するだけで様式変更に追従できる。
B2/B3/B10/B11/B12 等は仕様例に準拠。
"""

import os
from typing import Dict

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# (PropertyData キー) -> (テンプレートの行番号)。値は B 列、ラベルは A 列。
CELL_MAP = {
    "所在地": 2,
    "地番": 3,
    "家屋番号": 4,
    "地目": 5,
    "地積": 6,
    "種類": 7,
    "構造": 8,
    "床面積": 9,
    "用途地域": 10,
    "建ぺい率": 11,
    "容積率": 12,
    "防火地域": 13,
    "高度地区": 14,
    "洪水浸水想定": 15,
    "土砂災害": 16,
    "津波": 17,
    "最寄駅": 18,
    "駅距離": 19,
    "人口": 20,
    "世帯数": 21,
    "路線価": 22,
    "公示地価": 23,
    "所有者": 24,
    "抵当権": 25,
}

COMMENT_ROW = 27
TITLE = "重要事項説明書 ドラフト（調査支援システム出力）"


def ensure_template(template_path: str) -> None:
    """テンプレートが存在しなければ生成する。"""
    if os.path.exists(template_path):
        return
    os.makedirs(os.path.dirname(template_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "重説ドラフト"

    ws["A1"] = TITLE
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:B1")

    header_fill = PatternFill("solid", fgColor="D9E1F2")
    for key, row in CELL_MAP.items():
        cell = ws.cell(row=row, column=1, value=key)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        ws.cell(row=row, column=2, value="")

    ws.cell(row=COMMENT_ROW, column=1, value="AIコメント（下書き）").font = Font(bold=True)
    ws.cell(row=COMMENT_ROW, column=2, value="")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 60
    wb.save(template_path)


def export_excel(data: Dict[str, str], comment: str, template_path: str, output_path: str) -> str:
    """テンプレートに PropertyData を書き込み output_path に保存する。戻り値は出力パス。"""
    ensure_template(template_path)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = load_workbook(template_path)
    ws = wb.active

    for key, row in CELL_MAP.items():
        value = (data.get(key) or "").strip()
        ws.cell(row=row, column=2, value=value if value else "（要確認）")

    comment_cell = ws.cell(row=COMMENT_ROW, column=2, value=comment or "")
    comment_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[COMMENT_ROW].height = 90

    wb.save(output_path)
    return output_path
