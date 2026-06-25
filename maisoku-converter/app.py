"""他社マイソク → 自社 不動産案内書 Excel 変換ツール

Claude Code CLI を subprocess で呼び出して画像解析。
APIキー不要・Claude Pro/Max サブスクリプションで動作。
テンプレート XLS をそのまま利用し、値セルのみ書き換えて出力。
"""
from __future__ import annotations

import base64
import io
import json
import re
import subprocess
import tempfile
from pathlib import Path

import streamlit as st
import streamlit.components.v1 as components
from PIL import Image, ImageDraw
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string

# ── streamlit-drawable-canvas 互換パッチ ──────────────────────────────────────
# Streamlit 1.50 で image_to_url が移動されたため最小限の互換スタブを追加。
# ※ 背景画像は background_image パラメータを使わず init_drawing に直接埋め込む方式に
#   変更したため、このパッチが呼ばれることは基本的にない。
try:
    import streamlit.elements.image as _st_img
    if not hasattr(_st_img, "image_to_url"):
        def _image_to_url(image, width=-1, clamp=False, channels="RGB",
                          output_format="auto", image_id=""):
            if isinstance(image, str):
                return image
            try:
                from streamlit.runtime import get_instance, exists
                if exists():
                    buf = io.BytesIO()
                    if hasattr(image, "convert"):
                        image.convert("RGB").save(buf, "PNG")
                    return get_instance().media_file_mgr.add(
                        buf.getvalue(), "image/png", image_id)
            except Exception:
                pass
            return ""
        _st_img.image_to_url = _image_to_url
except Exception:
    pass

# ─── パス定数 ────────────────────────────────────────────────────────────────
APP_DIR      = Path(__file__).parent
LOGO_PATH    = APP_DIR / "company_logo.png"

# ─── 間取り図トレーサー（Gemini image-to-image）─────────────────────────────
try:
    from google import genai as _genai
    from google.genai import types as _gtypes
    _GEMINI_KEY = Path("/Users/apple/madori-tracer/config.py").read_text().split('"')[1]
    _gemini_client = _genai.Client(api_key=_GEMINI_KEY)
    TRACER_AVAILABLE = True
except Exception:
    TRACER_AVAILABLE = False

_TRACER_TYPE_RULES = {
    "マンション": """
- Emphasize corner structural pillars (マンション柱): solid filled black squares at all corners and wall junctions
- Include PS (パイプスペース) and MB (メーターボックス) labels
- Balcony (バルコニー) has thin border and white fill
- 浴室 and 洗面室 are always separate adjacent rooms — label both
""",
    "戸建て": """
- Include staircase (階段) with step lines if present
- Include entrance (玄関) with step/platform indication
- Garden/parking area (庭・駐車場) if shown: thin border, white fill, label
- No structural pillars (柱) unless explicitly shown
- Include 和室 (Japanese-style room) with tatami lines if present
""",
    "1K・1R": """
- Very simple layout: one main room + bathroom/kitchen unit
- Kitchen counter along wall: sink rectangle + stove circles
- Unit bath (ユニットバス): single rectangle with bathtub oval + toilet oval inside
- Compact layout — preserve exact proportions carefully
""",
    "その他": "",
}

_TRACER_BASE = """## RULE 1 — TRACING OPERATION (NOT DESIGN)
You must trace the original spatial layout exactly.
Every room, wall, door, and equipment item must stay in its EXACT original position and size.
Do NOT rearrange, move, add, or omit ANY element. Do NOT apply your own design preferences.

## RULE 2 — STYLE CONVERSION ONLY
Change ONLY the visual style: colors → black & white, simplified icons, clean fonts.
Spatial structure must be 100% identical to the input.

## RULE 3 — ROOM LABELING (MANDATORY)
Label EVERY room with its exact Japanese name from the original.
NEVER merge two separate rooms into one label.
- 浴室 and 洗面室 are ALWAYS separate rooms with separate labels
- 物入 and クローゼット are separate rooms
- All storage spaces must be individually labeled

## RULE 4 — EQUIPMENT POSITION (MANDATORY)
Place ALL equipment in the EXACT same position as the original.
Kitchen equipment (sink + stove) belongs ONLY inside the LDK/kitchen area — NEVER in bedrooms.
Do NOT relocate any equipment based on "typical" floor plan assumptions.

## RULE 5 — VISUAL STYLE
- Background: pure white
- Outer walls: thick solid black lines
- Interior partition walls: medium black lines
- All room fills: white (remove ALL color — no orange, tan, blue, gray, green)
- Room labels: copy EXACTLY as shown in the original image
  * If original shows name only (e.g. "玄関", "トイレ"): write just the name
  * If original shows name+size as one combined label (e.g. "洋室6帖"): write as ONE line — do NOT add a separate size line below
  * If original shows name and size on separate lines (e.g. "LDK" above "11.9帖"): write them on two separate lines
  * NEVER duplicate: if "6帖" is already inside the label, do NOT add "6帖" again below it
- Equipment icons (ultra-simplified):
  * Toilet: rectangle + oval bowl
  * Bathtub: rectangle + oval tub
  * Sink/洗面台: rectangle + circle
  * Kitchen: counter rectangle + sink box + 3 stove circles
  * Washing machine: rectangle + large circle
- Doors: straight line + quarter-circle arc (dashed)
- Windows: three parallel lines across wall
- PS / MB / 棚: keep as small text labels
- North compass: circle with filled black arrow + "N", bottom-right corner
"""


def _encode_image(image: Image.Image) -> bytes:
    img = image.copy()
    if img.width > 1400:
        r = 1400 / img.width
        img = img.resize((1400, int(img.height * r)), Image.LANCZOS)
    buf = io.BytesIO()
    img.convert("RGB").save(buf, "JPEG", quality=92)
    return buf.getvalue()


def _trace_madori(
    image: Image.Image,
    floor_type: str = "マンション",
    correction: str = "",
    prev_result: bytes | None = None,
) -> bytes:
    """間取り図画像をGeminiで白黒図面に引き直す。correction+prev_resultで再修正も可能。"""
    type_rules = _TRACER_TYPE_RULES.get(floor_type, "")

    if correction and prev_result:
        # 再修正：前回の白黒結果をベースに指定箇所だけ修正
        prev_img = Image.open(io.BytesIO(prev_result)).convert("RGB")
        img_bytes = _encode_image(prev_img)
        prompt = f"""This is a black-and-white floor plan that was already generated.
Make ONLY the following correction and leave everything else EXACTLY as-is:

"{correction}"

Do NOT redraw, restyle, or move anything that is not mentioned above.
Keep all room labels, wall positions, icons, and proportions identical to this image.
Output the corrected floor plan in the same black-and-white style."""

    elif correction:
        img_bytes = _encode_image(image)
        prompt = f"""Redraw this floor plan as a clean black-and-white floor plan.
Apply this correction: "{correction}"

{_TRACER_BASE}
## FLOOR TYPE: {floor_type}
{type_rules}"""

    else:
        img_bytes = _encode_image(image)
        prompt = f"""Redraw this floor plan as a clean black-and-white floor plan.

{_TRACER_BASE}
## FLOOR TYPE: {floor_type}
{type_rules}"""

    resp = _gemini_client.models.generate_content(
        model="gemini-3.1-flash-image",
        contents=[
            _gtypes.Part(inline_data=_gtypes.Blob(mime_type="image/jpeg", data=img_bytes)),
            _gtypes.Part(text=prompt),
        ],
        config=_gtypes.GenerateContentConfig(response_modalities=["IMAGE", "TEXT"]),
    )
    for part in resp.candidates[0].content.parts:
        if part.inline_data is not None:
            return part.inline_data.data
    texts = [p.text for p in resp.candidates[0].content.parts if p.text]
    raise RuntimeError("画像が生成されませんでした。\n" + "\n".join(texts)[:300])

# ── カスタムコンポーネント（画像配置エディタ）────────────────────────────────
_placement_editor = components.declare_component(
    "placement_editor",
    path=str(APP_DIR / "placement_component"),
)

def _pil_to_b64(img: Image.Image, max_size: int = 320) -> str:
    thumb = img.copy()
    thumb.thumbnail((max_size, max_size), Image.LANCZOS)
    buf = io.BytesIO()
    thumb.convert("RGB").save(buf, "JPEG", quality=82)
    return "data:image/jpeg;base64," + base64.b64encode(buf.getvalue()).decode()
COMPANY_JSON = APP_DIR / "company_info.json"

# テンプレート XLS（サンプルをそのまま使用）
TEMPLATE_XLS = {
    "賃貸": Path("/Users/apple/Downloads/ダイアパレス順慶町902号室.xls"),
    "売買": Path("/Users/apple/Downloads/加東市秋津売買.xls"),
}

# ─── 会社情報デフォルト ──────────────────────────────────────────────────────
COMPANY_DEFAULTS: dict = {
    "商号":           "大京商事株式会社",
    "宅建免許番号":    "大阪府知事（9）第27334",
    "所在地":         "大阪市都島区東野田町２－３－１４",
    "TEL":            "０６－６３５３－０４１８",
    "FAX":            "０６－６３５３－０２８０",
    "担当者":         "鷲見",
    "MAIL":           "info@daikyocorp.co.jp",
    "URL":            "http://www.daikyocorp.co.jp",
    "建設業免許番号":  "",
    "取引態様":       "売主",
    "チラシ":         "可",
    "情報誌":         "要確認",
    "インターネット": "要確認",
    "報酬形態":       "",
}

def load_company_info() -> dict:
    if COMPANY_JSON.exists():
        try:
            return json.loads(COMPANY_JSON.read_text(encoding="utf-8"))
        except Exception:
            pass
    return COMPANY_DEFAULTS.copy()

def save_company_info(info: dict) -> None:
    COMPANY_JSON.write_text(json.dumps(info, ensure_ascii=False, indent=2), encoding="utf-8")

# ─── Claude CLI ───────────────────────────────────────────────────────────────
CLAUDE_BIN     = "claude"
CLAUDE_TIMEOUT = 600

CLAUDE_PROMPT = """\
このファイルは不動産マイソク（物件チラシ）の画像です（ファイル名: {filename}）。
Read ツールでこのファイルを開いて内容を確認してください。
以下の JSON のみを返答してください（説明文・前置き一切不要）。

{{
  "template_type": "賃貸 または 売買",
  "specs": {{
    "物件名":     "物件名がなければ間取り＋所在地から作成",
    "種目":       "例: 分譲貸マンション / 中古戸建 / 売地",
    "賃料":       "賃貸のみ。例: 128000",
    "管理費":     "賃貸のみ。なければ空欄",
    "保証金":     "賃貸のみ",
    "敷金ヶ月":   "賃貸のみ。数字のみ: 1",
    "礼金ヶ月":   "賃貸のみ。数字のみ: 1",
    "更新料":     "賃貸のみ",
    "価格":       "売買のみ。例: 350",
    "所在地":     "住所",
    "交通":       "最寄り線と駅名。例: 地下鉄堺筋線・長堀鶴見緑地線 長堀橋駅",
    "交通2":      "徒歩分・バス等の補足。例: 地下鉄長堀鶴見緑地線 松屋町駅 徒歩5分",
    "物件名":     "建物名",
    "間取り":     "例: 1LDK",
    "間取り詳細": "例: 洋室10.5・LDK30.5",
    "構造":       "例: SRC / RC / 木",
    "階数_地上":  "数字のみ: 10",
    "部屋階":     "数字のみ: 9",
    "専有面積㎡": "数字のみ: 95.84",
    "専有面積坪": "数字のみ: 28.99",
    "バルコニー向き": "例: 南",
    "現況":       "例: 空家 / 居住中",
    "築年号":     "例: 昭和 / 平成",
    "築年":       "数字のみ: 57",
    "築月":       "数字のみ: 12",
    "引渡":       "例: 即時 / 相談",
    "設備":       "テキスト: 電気・都市ガス・上水道・公共下水道",
    "土地面積㎡": "売買のみ。数字のみ",
    "土地面積坪": "売買のみ。数字のみ",
    "延床面積㎡": "売買のみ。数字のみ",
    "延床面積坪": "売買のみ。数字のみ",
    "建ぺい率":   "売買のみ。数字のみ: 50",
    "容積率":     "売買のみ。数字のみ: 100",
    "その他":     ""
  }},
  "catchphrases": [
    "自社風キャッチコピー1（15文字以内）",
    "自社風キャッチコピー2（15文字以内）",
    "自社風キャッチコピー3（15文字以内）"
  ],
  "regions": [
    {{"種類": "外観写真", "x1": 0.0, "y1": 0.0, "x2": 0.0, "y2": 0.0}},
    {{"種類": "室内写真", "x1": 0.0, "y1": 0.0, "x2": 0.0, "y2": 0.0}},
    {{"種類": "間取り図", "x1": 0.0, "y1": 0.0, "x2": 0.0, "y2": 0.0}},
    {{"種類": "地図",     "x1": 0.0, "y1": 0.0, "x2": 0.0, "y2": 0.0}}
  ]
}}

regions: x1,y1=左上、x2,y2=右下（画像全体に対する割合 0.0〜1.0）
catchphrases: 元のコピーを使わず自社独自にリライト
"""

REGION_COLORS = {
    "外観写真": (220, 50,  50),
    "室内写真": (50,  120, 220),
    "間取り図": (50,  180, 80),
    "地図":    (160, 60,  200),
}

# ─── ユーティリティ ───────────────────────────────────────────────────────────

def convert_to_image(file_bytes: bytes, filename: str) -> Image.Image:
    ext = Path(filename).suffix.lower()
    if ext == ".pdf":
        try:
            import fitz
            doc = fitz.open(stream=file_bytes, filetype="pdf")
            pix = doc[0].get_pixmap(matrix=fitz.Matrix(2.0, 2.0))
            return Image.open(io.BytesIO(pix.tobytes("png"))).convert("RGB")
        except ImportError:
            raise RuntimeError("PDF変換には pymupdf が必要です: pip install pymupdf")
    return Image.open(io.BytesIO(file_bytes)).convert("RGB")


def analyze_with_claude(image: Image.Image) -> dict:
    analysis_img = image.copy()
    if analysis_img.width > 2000:
        r = 2000 / analysis_img.width
        analysis_img = analysis_img.resize((2000, int(analysis_img.height * r)), Image.LANCZOS)

    with tempfile.TemporaryDirectory(prefix="maisoku_") as tmp:
        p = Path(tmp) / "maisoku.png"
        analysis_img.save(p, "PNG")
        cmd = [CLAUDE_BIN, "-p", CLAUDE_PROMPT.format(filename=p.name),
               "--output-format", "json", "--tools", "Read",
               "--add-dir", tmp, "--dangerously-skip-permissions", "--model", "sonnet"]
        try:
            proc = subprocess.run(cmd, cwd=tmp, capture_output=True,
                                  text=True, timeout=CLAUDE_TIMEOUT)
        except FileNotFoundError:
            raise RuntimeError("`claude` コマンドが見つかりません。")
        except subprocess.TimeoutExpired:
            raise RuntimeError(f"解析タイムアウト（{CLAUDE_TIMEOUT}秒）。")

        if proc.returncode != 0:
            raise RuntimeError(f"Claude エラー (code={proc.returncode}):\n{proc.stderr[:400]}")

        try:
            result = json.loads(proc.stdout)
        except json.JSONDecodeError:
            raise RuntimeError("Claude レスポンスの JSON 解析失敗。")

        if result.get("is_error"):
            raise RuntimeError(f"Claude エラー: {result.get('result')}")

        raw = result.get("result", "")
        m = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", raw, re.DOTALL)
        json_str = m.group(1) if m else raw.strip()
        if not json_str.startswith("{"):
            m2 = re.search(r"\{.*\}", json_str, re.DOTALL)
            if m2:
                json_str = m2.group(0)
        return json.loads(json_str)


def crop_region(image: Image.Image, region: dict) -> Image.Image | None:
    x1, y1 = region.get("x1", 0), region.get("y1", 0)
    x2, y2 = region.get("x2", 0), region.get("y2", 0)
    if x2 <= x1 or y2 <= y1:
        return None
    w, h = image.size
    return image.crop((int(x1*w), int(y1*h), int(x2*w), int(y2*h)))


# ─── テンプレート XLS → openpyxl 変換 ───────────────────────────────────────

def _load_xls_as_openpyxl(xls_path: Path) -> tuple[openpyxl.Workbook,
                                                     openpyxl.worksheet.worksheet.Worksheet]:
    """XLS テンプレートを openpyxl Workbook として読み込む（構造・値を転写）。"""
    import xlrd

    xls  = xlrd.open_workbook(str(xls_path), formatting_info=True)
    xs   = xls.sheet_by_index(0)
    wb   = openpyxl.Workbook()
    ws   = wb.active
    ws.title = "案内書"

    # ページ設定
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins.left   = 0.20
    ws.page_margins.right  = 0.20
    ws.page_margins.top    = 0.30
    ws.page_margins.bottom = 0.30
    ws.page_margins.header = 0.10
    ws.page_margins.footer = 0.10

    # 列幅
    for ci in range(xs.ncols):
        cw      = xs.colinfo_map.get(ci)
        col_ltr = get_column_letter(ci + 1)
        if cw:
            ws.column_dimensions[col_ltr].width  = max(cw.width / 256, 0.1)
            ws.column_dimensions[col_ltr].hidden = bool(cw.hidden)

    # 行高さ
    for ri in range(xs.nrows):
        rh = xs.rowinfo_map.get(ri)
        if rh:
            ws.row_dimensions[ri + 1].height = rh.height / 20

    # マージセル（先にすべて解除してから再設定）
    for r1, r2, c1, c2 in xs.merged_cells:
        try:
            ws.merge_cells(
                start_row=r1 + 1, start_column=c1 + 1,
                end_row=r2,       end_column=c2,
            )
        except Exception:
            pass

    # xlrd line_style → openpyxl border style 変換テーブル
    _BORDER_STYLE = {
        1: "thin", 2: "medium", 3: "dashed", 4: "dotted",
        5: "thick", 6: "double", 7: "hair", 8: "mediumDashed",
        9: "dashDot", 10: "mediumDashDot", 11: "dashDotDot",
        12: "mediumDashDotDot", 13: "slantDashDot",
    }

    def _xlrd_color(idx: int) -> str:
        # 64 = automatic (black); それ以外はパレットから取得
        if idx == 64:
            return "000000"
        try:
            r, g, b = xls.colour_map.get(idx, (0, 0, 0))
            return f"{r:02X}{g:02X}{b:02X}"
        except Exception:
            return "000000"

    def _side(line_style: int, colour_index: int) -> Side | None:
        bs = _BORDER_STYLE.get(line_style)
        if bs is None:
            return None
        return Side(border_style=bs, color=_xlrd_color(colour_index))

    # セル値 & 罫線転写
    for r in range(xs.nrows):
        for c in range(xs.ncols):
            cell = ws.cell(r + 1, c + 1)
            v = xs.cell_value(r, c)
            if v != "":
                try:
                    cell.value = v
                    cell.font  = Font(name="ＭＳ Ｐゴシック", size=9)
                except Exception:
                    pass
            # 罫線
            try:
                xf_idx = xs.cell_xf_index(r, c)
                xf     = xls.xf_list[xf_idx]
                b      = xf.border
                left   = _side(b.left_line_style,   b.left_colour_index)
                right  = _side(b.right_line_style,  b.right_colour_index)
                top    = _side(b.top_line_style,    b.top_colour_index)
                bottom = _side(b.bottom_line_style, b.bottom_colour_index)
                if any(s is not None for s in (left, right, top, bottom)):
                    cell.border = Border(
                        left=left or Side(border_style=None),
                        right=right or Side(border_style=None),
                        top=top or Side(border_style=None),
                        bottom=bottom or Side(border_style=None),
                    )
            except Exception:
                pass

    # 印刷範囲（最初の 50 列 × 全行）
    last_vis = 50
    ws.print_area = f"A1:{get_column_letter(last_vis)}{xs.nrows}"

    return wb, ws


# ─── テンプレート値セル書き込み ──────────────────────────────────────────────

def _w(ws, row: int, col: str | int, value, sz: int = 9, bold: bool = False):
    """マージ済みでも master cell に値を書き込む。"""
    col_idx = column_index_from_string(col) if isinstance(col, str) else col
    try:
        cell = ws.cell(row, col_idx)
        cell.value = value
        cell.font  = Font(name="ＭＳ Ｐゴシック", size=sz, bold=bold)
    except Exception:
        pass


def _fill_chintai(ws, specs: dict, company: dict, catchphrases: list[str]):
    """賃貸テンプレート（ダイアパレス形式）の値セルを書き換える。
    テンプレート: ダイアパレス順慶町902号室.xls
    """
    # ── キャッチコピー (L1:AJ2 エリア、テンプレートは空き) ─────────────────
    cp = "　".join(f"◆ {c}" for c in catchphrases if c)
    _w(ws, 1, "L", cp, sz=9)

    # ── 種目 (AL3:AX4 merged, master=AL3) ────────────────────────────────────
    _w(ws, 3, "AL", specs.get("種目", ""), sz=9)

    # ── 賃料 (AN5:AV7 merged, master=AN5) ────────────────────────────────────
    rent = specs.get("賃料", "")
    try:
        _w(ws, 5, "AN", int(re.sub(r"[^0-9]", "", str(rent))) if rent else "")
    except ValueError:
        _w(ws, 5, "AN", rent)

    # ── 管理費 (AL9:AM9, '込' セルを上書き or 金額) ──────────────────────────
    kanrihi = specs.get("管理費", "")
    if kanrihi:
        _w(ws, 9, "AM", kanrihi)

    # ── 保証金 / 敷金 / 礼金 ─────────────────────────────────────────────────
    _w(ws, 10, "AW", specs.get("敷金ヶ月", ""))
    _w(ws, 11, "AW", specs.get("礼金ヶ月", ""))
    # ヶ月ラベル列が狭いため幅を補正
    ws.column_dimensions['AQ'].width = 5.0
    ws.column_dimensions['AX'].width = 5.0

    # ── 所在地 (AL13:AX14 merged, master=AL13) ───────────────────────────────
    _w(ws, 13, "AL", specs.get("所在地", ""), sz=9)

    # ── 交通 (AL17:AX18 merged, master=AL17) ─────────────────────────────────
    _w(ws, 17, "AL", specs.get("交通", ""), sz=9)
    # 他交通手段 (AO20:AX21 merged, master=AO20)
    _w(ws, 20, "AO", specs.get("交通2", ""), sz=9)

    # ── 建物名 (AO22:AX24 merged, master=AO22) ───────────────────────────────
    _w(ws, 22, "AO", specs.get("物件名", ""), sz=9)

    # ── 構造 (AO25:AS25 merged, master=AO25) ─────────────────────────────────
    _w(ws, 25, "AO", specs.get("構造", ""))
    _w(ws, 25, "AS", specs.get("階数_地上", ""))   # 地上N階
    # 部屋階 (AP26:AQ26, master=AP26)
    _w(ws, 26, "AP", specs.get("部屋階", ""))

    # ── 間取り (AO27, AT27:AX27 詳細) ─────────────────────────────────────────
    _w(ws, 27, "AO", specs.get("間取り", ""), sz=9)
    _w(ws, 27, "AT", specs.get("間取り詳細", ""), sz=8)

    # ── 面積 (AQ29 ㎡, AV29 坪) ──────────────────────────────────────────────
    _w(ws, 29, "AQ", specs.get("専有面積㎡", ""))
    _w(ws, 29, "AV", specs.get("専有面積坪", ""))

    # ── バルコニー向き (AO31) ─────────────────────────────────────────────────
    _w(ws, 31, "AO", specs.get("バルコニー向き", ""))

    # ── 現況 (AV33) ──────────────────────────────────────────────────────────
    _w(ws, 33, "AV", specs.get("現況", ""))

    # ── 建築 (AO34=年号, AQ34=年, AT34=月) ───────────────────────────────────
    _w(ws, 34, "AO", specs.get("築年号", "昭和"))
    try:
        _w(ws, 34, "AQ", int(specs.get("築年", 0)) or "")
    except Exception:
        _w(ws, 34, "AQ", specs.get("築年", ""))
    try:
        _w(ws, 34, "AT", int(specs.get("築月", 0)) or "")
    except Exception:
        _w(ws, 34, "AT", specs.get("築月", ""))

    # ── 引渡 (AO44) ──────────────────────────────────────────────────────────
    _w(ws, 44, "AO", specs.get("引渡", ""))

    # ── 設備 (AL46:AX51 merged, master=AL46) ─────────────────────────────────
    _w(ws, 46, "AL", specs.get("設備", ""), sz=8)

    # ── 会社情報 (rows 53-58) ─────────────────────────────────────────────────
    _w(ws, 53, "H",  company.get("商号", ""),          sz=16, bold=True)
    _w(ws, 53, "Z",  company.get("所在地", ""),         sz=8)
    _w(ws, 53, "AM", company.get("チラシ", ""),         sz=8)
    _w(ws, 53, "AP", company.get("取引態様", ""),       sz=9)
    _w(ws, 53, "AT", company.get("報酬形態", ""),       sz=9)
    _w(ws, 55, "Z",  company.get("TEL", ""),            sz=9)
    _w(ws, 55, "AM", company.get("情報誌", ""),         sz=8)
    _w(ws, 56, "Z",  company.get("FAX", ""),            sz=9)
    _w(ws, 57, "K",  company.get("宅建免許番号", ""),   sz=8)
    _w(ws, 57, "U",  company.get("担当者", ""),         sz=9)
    _w(ws, 57, "Z",  company.get("MAIL", ""),           sz=8)
    _w(ws, 57, "AM", company.get("インターネット", ""), sz=8)
    _w(ws, 58, "K",  company.get("建設業免許番号", ""),   sz=8)
    _w(ws, 58, "Z",  company.get("URL", ""),            sz=8)


def _fill_baibai(ws, specs: dict, company: dict, catchphrases: list[str]):
    """売買テンプレート（加東市秋津形式）の値セルを書き換える。
    テンプレート: 加東市秋津売買.xls
    """
    cp = "　".join(f"◆ {c}" for c in catchphrases if c)
    _w(ws, 1, "L", cp, sz=9)

    # 種目 (AK3:AX4 merged, master=AK3)
    _w(ws, 3, "AK", specs.get("種目", ""), sz=9)

    # 価格 (AN5:AT7 merged, master=AN5)
    price = specs.get("価格", "")
    try:
        _w(ws, 5, "AN", float(re.sub(r"[^0-9.]", "", str(price))) if price else "")
    except ValueError:
        _w(ws, 5, "AN", price)

    # 所在地 (AK9:AX12 master=AK9)
    _w(ws, 9, "AK", specs.get("所在地", ""), sz=9)

    # 交通 (AK13:AX14 master=AK13)
    _w(ws, 13, "AK", specs.get("交通", ""), sz=9)
    _w(ws, 16, "AK", specs.get("交通2", ""), sz=8)

    # 土地面積 (AP18 ㎡, AU18 坪)
    _w(ws, 18, "AP", specs.get("土地面積㎡", ""))
    _w(ws, 18, "AU", specs.get("土地面積坪", ""))

    # 建物 - 構造 (AN35)
    _w(ws, 35, "AN", specs.get("構造", ""))
    _w(ws, 35, "AS", specs.get("階数_地上", ""))

    # 間取り (AN37)
    _w(ws, 37, "AN", specs.get("間取り", ""))

    # 延床面積 (AN39 ㎡, AS39 坪)
    _w(ws, 39, "AN", specs.get("延床面積㎡", ""))
    _w(ws, 39, "AS", specs.get("延床面積坪", ""))

    # 建築 (AP41=年, AS41=月)
    try:
        _w(ws, 41, "AP", int(specs.get("築年", 0)) or "")
    except Exception:
        _w(ws, 41, "AP", specs.get("築年", ""))
    try:
        _w(ws, 41, "AS", int(specs.get("築月", 0)) or "")
    except Exception:
        _w(ws, 41, "AS", specs.get("築月", ""))

    # 現況 (AN43)
    _w(ws, 43, "AN", specs.get("現況", ""))

    # 引渡 (AN44)
    _w(ws, 44, "AN", specs.get("引渡", ""))

    # 設備 (AK46 テキスト)
    _w(ws, 46, "AK", specs.get("設備", ""), sz=8)

    # 会社情報 (rows 50-55)
    _w(ws, 50, "H",  company.get("商号", ""),          sz=16, bold=True)
    _w(ws, 50, "Z",  company.get("所在地", ""),         sz=8)
    _w(ws, 50, "AL", company.get("チラシ", ""),         sz=8)
    _w(ws, 50, "AO", company.get("取引態様", ""),       sz=9)
    _w(ws, 52, "Z",  company.get("TEL", ""),            sz=9)
    _w(ws, 52, "AL", company.get("情報誌", ""),         sz=8)
    _w(ws, 53, "Z",  company.get("FAX", ""),            sz=9)
    _w(ws, 54, "K",  company.get("宅建免許番号", ""),   sz=8)
    _w(ws, 54, "U",  company.get("担当者", ""),         sz=9)
    _w(ws, 54, "Z",  company.get("MAIL", ""),           sz=8)
    _w(ws, 54, "AL", company.get("インターネット", ""), sz=8)
    _w(ws, 55, "K",  company.get("建設業免許番号", ""),   sz=8)
    _w(ws, 55, "Z",  company.get("URL", ""),            sz=8)


# ─── 画像挿入 ─────────────────────────────────────────────────────────────────

def _insert_image(ws, pil_img: Image.Image,
                  col_from: int, row_from: int,
                  col_to: int,   row_to: int,
                  col_w: float = 3.89, row_h: float = 12.0):
    """TwoCellAnchor で画像を指定セル範囲に収まるよう挿入する。"""
    px_per_char = 7.0
    # 1pt = 96/72 px (96dpi)。int() の丸め誤差を防ぐため round() を使用
    w_px = round((col_to - col_from) * col_w * px_per_char)
    h_px = round((row_to - row_from) * row_h * (96 / 72))

    img = pil_img.convert("RGB").copy()
    img.thumbnail((w_px, h_px), Image.LANCZOS)
    # 白背景にセンタリング
    canvas = Image.new("RGB", (w_px, h_px), (255, 255, 255))
    canvas.paste(img, ((w_px - img.width) // 2, (h_px - img.height) // 2))

    buf = io.BytesIO()
    # 96dpi を明示して保存（指定なしだと LibreOffice/Excel が 72dpi と解釈し133%に拡大する）
    canvas.save(buf, "PNG", dpi=(96, 96))
    buf.seek(0)
    xl = XLImage(buf)
    xl.width  = w_px
    xl.height = h_px

    anchor = TwoCellAnchor()
    anchor._from = AnchorMarker(col=col_from - 1, colOff=0,
                                row=row_from - 1, rowOff=0)
    anchor.to    = AnchorMarker(col=col_to - 1,   colOff=0,
                                row=row_to - 1,   rowOff=0)
    anchor.editAs = "twoCell"
    xl.anchor = anchor
    ws.add_image(xl)


# ─── メイン Excel 生成 ────────────────────────────────────────────────────────

def create_fudosan_excel(
    template_type: str,
    specs: dict,
    company_info: dict,
    catchphrases: list[str],
    image_placements: list[dict],   # [{"img": PIL, "c1": int, "r1": int, "c2": int, "r2": int}, ...]
) -> bytes:
    """テンプレート XLS を読み込み、値セル・画像を書き換えて XLSX を返す。"""
    is_chintai = template_type != "売買"
    tpl_key    = "賃貸" if is_chintai else "売買"
    tpl_path   = TEMPLATE_XLS.get(tpl_key)

    if tpl_path is None or not tpl_path.exists():
        raise RuntimeError(f"テンプレート XLS が見つかりません: {tpl_path}")

    wb, ws = _load_xls_as_openpyxl(tpl_path)

    if is_chintai:
        _fill_chintai(ws, specs, company_info, catchphrases)
        col_w, row_h = 3.89, 12.0
    else:
        _fill_baibai(ws, specs, company_info, catchphrases)
        col_w, row_h = 3.66, 13.3

    # 画像挿入
    for pl in image_placements:
        pil = pl.get("img")
        if pil is None:
            continue
        try:
            _insert_image(ws, pil,
                          pl["c1"], pl["r1"],
                          pl["c2"], pl["r2"],
                          col_w=col_w, row_h=row_h)
        except Exception as e:
            pass  # 画像配置失敗は無視して続行

    # ロゴ挿入（会社情報帯の左枠）
    if LOGO_PATH.exists():
        try:
            logo_pil = Image.open(LOGO_PATH).convert("RGBA")
            bg = Image.new("RGB", logo_pil.size, (255, 255, 255))
            bg.paste(logo_pil, mask=logo_pil.split()[3])
            if is_chintai:
                _insert_image(ws, bg, 1, 53, 7, 59, col_w=col_w, row_h=row_h)
            else:
                _insert_image(ws, bg, 1, 50, 7, 56, col_w=col_w, row_h=row_h)
        except Exception:
            pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─── Streamlit UI ─────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="不動産案内書 変換ツール",
    layout="wide",
    page_icon="🏠",
    initial_sidebar_state="expanded",
)

# ── サイドバー: 会社情報 ─────────────────────────────────────────────────────
with st.sidebar:
    st.header("🏢 自社情報設定")
    st.caption("入力内容は「保存」ボタンで記録されます")
    company = load_company_info()
    ci: dict = {}

    def _si(label, key, default=""):
        return st.text_input(label, value=company.get(key, default))

    ci["商号"]           = _si("商号（会社名）",   "商号")
    ci["宅建免許番号"]    = _si("宅建免許番号",     "宅建免許番号")
    ci["建設業免許番号"]  = _si("建設業免許番号",   "建設業免許番号")
    ci["所在地"]         = _si("所在地",           "所在地")
    ci["TEL"]            = _si("TEL",              "TEL")
    ci["FAX"]            = _si("FAX",              "FAX")
    ci["担当者"]         = _si("担当者",           "担当者")
    ci["MAIL"]           = _si("E-MAIL",           "MAIL")
    ci["URL"]            = _si("URL",              "URL")

    def _ss(label, key, opts):
        cur = company.get(key, opts[0])
        return st.selectbox(label, opts, index=opts.index(cur) if cur in opts else 0)

    ci["取引態様"]       = _ss("取引態様",       "取引態様",       ["売主","貸主","代理","仲介","専任","一般"])
    ci["チラシ"]         = _ss("チラシ",         "チラシ",         ["可","不可","要確認"])
    ci["情報誌"]         = _ss("情報誌",         "情報誌",         ["可","不可","要確認"])
    ci["インターネット"] = _ss("インターネット", "インターネット", ["可","不可","要確認"])
    ci["報酬形態"]       = st.text_input("報酬形態", value=company.get("報酬形態", ""))

    if st.button("💾 自社情報を保存", type="primary", use_container_width=True):
        save_company_info(ci)
        st.success("✅ 保存しました")
        company = ci

    st.divider()
    st.caption(f"保存先: `{COMPANY_JSON.name}`")

    st.divider()
    st.subheader("🖼️ 会社ロゴ")
    st.caption("案内書下帯の左枠に挿入されます")
    if LOGO_PATH.exists():
        st.image(str(LOGO_PATH), use_container_width=True)
        st.caption("現在のロゴ")
    logo_upload = st.file_uploader("ロゴ画像をアップロード", type=["png","jpg","jpeg"], key="logo_upload")
    if logo_upload:
        try:
            logo_pil = Image.open(logo_upload).convert("RGBA")
            logo_pil.save(LOGO_PATH, "PNG")
            st.success("✅ ロゴを保存しました")
            st.rerun()
        except Exception as e:
            st.error(f"ロゴ保存エラー: {e}")

# ── メインエリア ──────────────────────────────────────────────────────────────
st.title("🏠 他社マイソク → 不動産案内書 Excel 変換ツール")
st.caption("他社チラシ（PDF・画像）をアップロードすると Claude AI が解析し、"
           "自社テンプレートに値を書き込んだ Excel を出力します。")

uploaded = st.file_uploader(
    "他社マイソクをアップロード（PDF・PNG・JPG対応）",
    type=["pdf", "png", "jpg", "jpeg"],
)

if not uploaded:
    st.info("↑ まずファイルをアップロードしてください")
    st.stop()

col_left, col_right = st.columns([1, 2])
with col_left:
    st.subheader("元のマイソク")
    try:
        original_image = convert_to_image(uploaded.getvalue(), uploaded.name)
        st.image(original_image, use_container_width=True)
    except Exception as e:
        st.error(f"プレビューエラー: {e}")
        st.stop()

with col_right:
    st.subheader("AI 解析")
    if st.button("🤖 AI で解析開始", type="primary", use_container_width=True):
        with st.spinner("Claude が解析中です（最大 10 分）..."):
            try:
                result = analyze_with_claude(original_image)
                st.session_state["analysis"]       = result
                st.session_state["original_image"] = original_image
                st.success("✅ 解析完了！下のタブで内容を確認・編集してください。")
                st.rerun()
            except Exception as e:
                st.error(f"❌ 解析エラー: {e}")

if "analysis" not in st.session_state:
    st.info("「AI で解析開始」ボタンを押してください。")
    st.stop()

analysis      = st.session_state["analysis"]
orig_img      = st.session_state.get("original_image", original_image)
specs         = analysis.get("specs", {})
catchphrases  = analysis.get("catchphrases", ["", "", ""])
regions       = analysis.get("regions", [])
detected_type = analysis.get("template_type", "賃貸")

st.divider()
st.subheader("📝 内容確認・編集")
tab_spec, tab_cp, tab_img = st.tabs(["物件スペック", "キャッチコピー", "切り抜き・配置"])

# ── スペック編集 ──────────────────────────────────────────────────────────────
with tab_spec:
    col_type, _ = st.columns([1, 3])
    with col_type:
        tmpl = st.radio("書類種別", ["賃貸", "売買"],
                        index=0 if "売買" not in detected_type else 1,
                        horizontal=True)

    edited_specs: dict = {}
    c1, c2 = st.columns(2)

    if tmpl == "賃貸":
        lf = ["物件名","種目","賃料","管理費","保証金","敷金ヶ月","礼金ヶ月","更新料"]
        rf = ["所在地","交通","交通2","間取り","間取り詳細","専有面積㎡","専有面積坪",
              "構造","階数_地上","部屋階","バルコニー向き","築年号","築年","築月","現況","引渡"]
    else:
        lf = ["物件名","種目","価格","土地面積㎡","土地面積坪","建ぺい率","容積率"]
        rf = ["所在地","交通","交通2","間取り","延床面積㎡","延床面積坪",
              "構造","階数_地上","築年","築月","現況","引渡"]

    with c1:
        for f in lf:
            edited_specs[f] = st.text_input(f, value=str(specs.get(f, "")), key=f"s_{f}")
    with c2:
        for f in rf:
            edited_specs[f] = st.text_input(f, value=str(specs.get(f, "")), key=f"s_{f}")

    edited_specs["設備"]  = st.text_area("設備（設備一覧テキスト）",
                                         value=str(specs.get("設備", "")), height=70, key="s_setubi")
    edited_specs["その他"] = st.text_area("その他",
                                          value=str(specs.get("その他", "")), height=50, key="s_other")

# ── キャッチコピー編集 ────────────────────────────────────────────────────────
with tab_cp:
    edited_cps: list[str] = []
    for i in range(3):
        default = catchphrases[i] if i < len(catchphrases) else ""
        edited_cps.append(st.text_input(f"キャッチコピー {i+1}", value=default, key=f"cp_{i}"))

# ── 切り抜き・配置 ────────────────────────────────────────────────────────────
with tab_img:
    all_regions    = [r for r in regions if r.get("種類")]
    adjusted_regions: list[dict] = []

    # ── インタラクティブ枠編集（streamlit-drawable-canvas）──────────────────
    st.markdown("**🎯 領域を調整** — 枠をクリック選択 → ドラッグで移動 / 角をドラッグでリサイズ")

    try:
        from streamlit_drawable_canvas import st_canvas

        img_w, img_h = orig_img.size
        canvas_w = min(720, img_w)
        canvas_h = int(img_h * canvas_w / img_w)

        canvas_objects: list[dict] = []
        for reg in all_regions:
            kind  = reg.get("種類", "不明")
            color = REGION_COLORS.get(kind, (150, 150, 150))
            bx    = reg.get("x1", 0) * canvas_w
            by    = reg.get("y1", 0) * canvas_h
            bw    = (reg.get("x2", 0) - reg.get("x1", 0)) * canvas_w
            bh    = (reg.get("y2", 0) - reg.get("y1", 0)) * canvas_h
            if bw > 2 and bh > 2:
                canvas_objects.append({
                    "type": "rect",
                    "name": kind,   # 種類を name に保存 → 並び替えに依存しない
                    "left": round(bx, 1), "top": round(by, 1),
                    "width": round(bw, 1), "height": round(bh, 1),
                    "fill":   "rgba({},{},{},0.20)".format(*color),
                    "stroke": "rgb({},{},{})".format(*color),
                    "strokeWidth": 2, "selectable": True,
                })

        # 下絵を base64 JPEG に変換して fabric.js Image オブジェクトとして埋め込む
        # ※ background_image パラメータは image_to_url 経由のためStreamlit 1.50 で
        #   動作しない。init_drawing に直接入れることで image_to_url を完全に迂回する。
        _buf = io.BytesIO()
        _bg = orig_img.resize((canvas_w, canvas_h), Image.LANCZOS).convert("RGB")
        _bg.save(_buf, "JPEG", quality=65)
        _bg_b64 = "data:image/jpeg;base64," + base64.b64encode(_buf.getvalue()).decode()
        bg_fabric_obj = {
            "type": "image", "src": _bg_b64,
            "left": 0, "top": 0, "scaleX": 1.0, "scaleY": 1.0,
            "selectable": False, "evented": False,
            "lockMovementX": True, "lockMovementY": True,
            "lockScalingX": True, "lockScalingY": True,
            "hasControls": False, "hasBorders": False,
            "name": "_background_",
        }
        init_drawing = {
            "version": "4.4.0",
            "objects": [bg_fabric_obj] + canvas_objects,
        }

        canvas_result = st_canvas(
            fill_color       = "rgba(255,255,255,0.0)",
            stroke_width     = 2,
            background_color = "",
            update_streamlit = True,
            height           = canvas_h,
            width            = canvas_w,
            drawing_mode     = "transform",
            initial_drawing  = init_drawing,
            display_toolbar  = False,
            key              = "region_canvas",
        )

        # stroke色 → 種類の逆引き辞書（fabric.jsはnameを保存しないがstroke色は保存する）
        _stroke_to_kind = {
            "rgb({},{},{})".format(*v): k for k, v in REGION_COLORS.items()
        }

        # キャンバス結果から座標を取得
        if canvas_result.json_data and canvas_result.json_data.get("objects"):
            _obj_list = [o for o in canvas_result.json_data["objects"]
                         if o.get("name") != "_background_" and o.get("type") == "rect"]
            for idx, obj in enumerate(_obj_list):
                # stroke色で種類を特定。色が取れない場合は元のall_regionsの順番で補完
                stroke = obj.get("stroke", "")
                kind = _stroke_to_kind.get(stroke)
                if kind is None:
                    kind = all_regions[idx]["種類"] if idx < len(all_regions) else "不明"
                sx  = obj.get("scaleX", 1.0)
                sy  = obj.get("scaleY", 1.0)
                lft = obj.get("left", 0)
                top = obj.get("top",  0)
                ow  = obj.get("width",  0) * sx
                oh  = obj.get("height", 0) * sy
                adjusted_regions.append({
                    "種類": kind,
                    "x1": max(0.0, min(1.0, lft / canvas_w)),
                    "y1": max(0.0, min(1.0, top / canvas_h)),
                    "x2": max(0.0, min(1.0, (lft + ow) / canvas_w)),
                    "y2": max(0.0, min(1.0, (top + oh) / canvas_h)),
                })
        else:
            adjusted_regions = [
                r.copy() for r in all_regions
                if r.get("x2", 0) > r.get("x1", 0) and r.get("y2", 0) > r.get("y1", 0)
            ]

        # 座標リードアウト
        if adjusted_regions:
            coord_cols = st.columns(min(len(adjusted_regions), 4))
            for i, reg in enumerate(adjusted_regions):
                kind = reg["種類"]
                c    = REGION_COLORS.get(kind, (150, 150, 150))
                hx   = "#{:02X}{:02X}{:02X}".format(*c)
                with coord_cols[i % len(coord_cols)]:
                    st.markdown(
                        f"<b style='color:{hx}'>■ {kind}</b><br>"
                        f"x: {reg['x1']:.2f}–{reg['x2']:.2f}<br>"
                        f"y: {reg['y1']:.2f}–{reg['y2']:.2f}",
                        unsafe_allow_html=True,
                    )

    except ImportError:
        st.warning("streamlit-drawable-canvas が未インストール。スライダーで代替中。")
        for i, reg in enumerate(all_regions):
            kind  = reg.get("種類", "不明")
            color = REGION_COLORS.get(kind, (150, 150, 150))
            with st.expander(f"**{kind}** を調整する", expanded=(i == 0)):
                c_sl, c_pv = st.columns(2)
                with c_sl:
                    x1 = st.slider("← 左端", 0.0, 1.0, float(reg.get("x1", 0.0)), 0.01, key=f"adj_{i}_x1")
                    x2 = st.slider("→ 右端", 0.0, 1.0, float(reg.get("x2", 1.0)), 0.01, key=f"adj_{i}_x2")
                    y1 = st.slider("↑ 上端", 0.0, 1.0, float(reg.get("y1", 0.0)), 0.01, key=f"adj_{i}_y1")
                    y2 = st.slider("↓ 下端", 0.0, 1.0, float(reg.get("y2", 1.0)), 0.01, key=f"adj_{i}_y2")
                adj = {"種類": kind, "x1": x1, "y1": y1, "x2": x2, "y2": y2}
                with c_pv:
                    if x2 > x1 and y2 > y1:
                        pv = crop_region(orig_img, adj)
                        if pv:
                            st.image(pv, use_container_width=True)
                adjusted_regions.append(adj)

    # ── 切り抜きプレビュー ────────────────────────────────────────────────────
    st.divider()
    cropped: dict[str, Image.Image] = {}
    valid_adj = [r for r in adjusted_regions
                 if r.get("x2", 0) > r.get("x1", 0) and r.get("y2", 0) > r.get("y1", 0)]

    for reg in valid_adj:
        kind = reg["種類"]
        img  = crop_region(orig_img, reg)
        if img:
            key = (kind if kind not in cropped
                   else f"{kind}_{sum(1 for k in cropped if k.startswith(kind))+1}")
            cropped[key] = img

    # 清書済み画像を適用（rerun後も維持）
    _traced_img: Image.Image | None = st.session_state.get("madori_traced_img")
    _traced_key: str | None = st.session_state.get("madori_traced_key")
    if _traced_img is not None and _traced_key is not None and _traced_key in cropped:
        cropped[_traced_key] = _traced_img

    # プレビュー表示
    if cropped:
        prev_cols = st.columns(min(len(cropped), 4))
        for i, (kind, img) in enumerate(cropped.items()):
            c  = REGION_COLORS.get(kind, (150, 150, 150))
            hx = "#{:02X}{:02X}{:02X}".format(*c)
            with prev_cols[i % len(prev_cols)]:
                label = f"<span style='color:{hx}'>■ **{kind}**</span>"
                if kind == _traced_key and _traced_img is not None:
                    label += " ✅ 清書済み"
                st.markdown(label, unsafe_allow_html=True)
                st.image(img, use_container_width=True)

    # ── 間取り図 AI清書（間取り図トレーサー連携）────────────────────────────
    if cropped and TRACER_AVAILABLE:
        st.divider()
        st.markdown("**🏠 間取り図をAI清書**")

        _traced_key: str | None = st.session_state.get("madori_traced_key")
        _is_traced = _traced_key is not None and _traced_key in cropped

        # 対象画像を選択
        _crop_labels = list(cropped.keys())
        _default_idx = (
            _crop_labels.index(_traced_key) if _is_traced and _traced_key in _crop_labels
            else next((i for i, k in enumerate(_crop_labels) if "間取" in k), 0)
        )
        _ta, _tb = st.columns([3, 1])
        with _ta:
            _target_label = st.radio(
                "清書する画像を選択",
                _crop_labels,
                index=_default_idx,
                horizontal=True,
                key="tracer_target",
            )
        with _tb:
            _floor_type = st.selectbox(
                "図面タイプ", ["マンション", "戸建て", "1K・1R", "その他"],
                key="tracer_floor_type",
            )

        _t1, _t2, _t3 = st.columns([1, 1, 1])
        with _t1:
            _do_trace = st.button(
                "✨ AI清書実行" if not _is_traced else "🔄 再清書",
                key="btn_trace", use_container_width=True, type="primary",
            )
        with _t2:
            if _is_traced and st.button("↩ 元の切り抜きに戻す", key="btn_reset_trace", use_container_width=True):
                for _k in ["madori_traced_img", "madori_traced_bytes", "madori_traced_key"]:
                    st.session_state.pop(_k, None)
                st.rerun()

        # 清書済みの場合はbefore/after表示
        if _is_traced:
            _raw_for_show = crop_region(orig_img, next(
                (r for r in valid_adj if r.get("種類") == _traced_key), valid_adj[0]
            )) if valid_adj else cropped[_traced_key]
            _bc1, _bc2 = st.columns(2)
            with _bc1:
                st.caption(f"元の切り抜き（{_traced_key}）")
                st.image(_raw_for_show, use_container_width=True)
            with _bc2:
                st.caption("AI清書後 ✅")
                st.image(cropped[_traced_key], use_container_width=True)

        # 再修正テキスト入力（清書済みの場合のみ表示）
        _correction_text = ""
        if _is_traced:
            _correction_text = st.text_input(
                "修正指示（任意）",
                placeholder="例: 浴室と洗面室のラベルを分けて / 北コンパスを追加して",
                key="tracer_correction",
            )

        if _do_trace:
            # 元の切り抜き（生画像）を取得
            _target_reg = next(
                (r for r in valid_adj if r.get("種類") == _target_label), None
            )
            _raw_crop = crop_region(orig_img, _target_reg) if _target_reg else cropped.get(_target_label)
            if _raw_crop is None:
                st.error("対象画像が見つかりません。")
            else:
                # 再修正の場合は前回結果バイトをprev_resultとして渡す
                _prev_bytes = st.session_state.get("madori_traced_bytes") if _correction_text else None
                _spin_msg = (
                    f"修正指示で再生成中...（30秒ほど）" if _correction_text
                    else f"Geminiが「{_target_label}」を引き直し中...（30秒ほど）"
                )
                with st.spinner(_spin_msg):
                    try:
                        _result_bytes = _trace_madori(
                            _raw_crop, _floor_type,
                            correction=_correction_text,
                            prev_result=_prev_bytes,
                        )
                        _result_img = Image.open(io.BytesIO(_result_bytes)).convert("RGB")
                        st.session_state["madori_traced_img"]   = _result_img
                        st.session_state["madori_traced_bytes"] = _result_bytes
                        st.session_state["madori_traced_key"]   = _target_label
                        st.rerun()
                    except Exception as _e:
                        st.error(f"❌ 清書エラー: {_e}")

    elif cropped and not TRACER_AVAILABLE:
        st.divider()
        st.info("💡 間取り図AI清書: `/Users/apple/madori-tracer/config.py` にGemini APIキーを設定すると利用できます。")

    # ── 手動差し替え ──────────────────────────────────────────────────────────
    st.divider()
    with st.expander("✋ 手動アップロードで差し替え（任意）"):
        mu1, mu2, mu3, mu4 = st.columns(4)
        with mu1: mf1 = st.file_uploader("外観写真", type=["png","jpg","jpeg"], key="mf1")
        with mu2: mf2 = st.file_uploader("室内写真", type=["png","jpg","jpeg"], key="mf2")
        with mu3: mf3 = st.file_uploader("間取り図", type=["png","jpg","jpeg"], key="mf3")
        with mu4: mf4 = st.file_uploader("地図",     type=["png","jpg","jpeg"], key="mf4")

    for lbl, mf in {"外観写真": mf1, "室内写真": mf2, "間取り図": mf3, "地図": mf4}.items():
        if mf:
            cropped[lbl] = Image.open(mf).convert("RGB")

    # ── 画像配置プレビュー（インタラクティブ）────────────────────────────────
    st.divider()
    st.markdown("**🖼️ Excel 画像配置プレビュー**")
    st.caption("チップをクリックで使用/非使用切り替え｜ドラッグで移動｜右下角をドラッグでリサイズ")

    # テンプレートの画像エリア（賃貸: cols A-AJ = 1-36, rows 3-52）
    IMG_COL_START = 1
    IMG_COL_END   = 36 if tmpl == "賃貸" else 35
    IMG_ROW_START = 3
    IMG_ROW_END   = 52 if tmpl == "賃貸" else 48

    # プレビューキャンバス（A4横比率）
    PREV_W = 600
    PREV_H = int(PREV_W * (210 / 297))  # A4横比率 ≈ 424

    total_cols     = 50
    img_area_ratio = IMG_COL_END / total_cols  # ~0.72

    placements: list[dict] = []

    if cropped:
        images_b64 = {label: _pil_to_b64(img) for label, img in cropped.items()}
        init_pl    = st.session_state.get("placement_data", {})

        placement_result = _placement_editor(
            images           = images_b64,
            container_w      = PREV_W,
            container_h      = PREV_H,
            img_area_ratio   = img_area_ratio,
            initial_placements = init_pl,
            key              = "placement_editor",
            default          = None,
        )

        if placement_result:
            st.session_state["placement_data"] = placement_result

        # placement_result を Excel セル座標に変換
        pd_now   = placement_result or st.session_state.get("placement_data") or {}
        img_cols = IMG_COL_END - IMG_COL_START + 1
        img_rows = IMG_ROW_END - IMG_ROW_START + 1
        IA_W     = int(PREV_W * img_area_ratio)  # 画像エリアのピクセル幅

        for label, pl in pd_now.items():
            if not pl.get("enabled", True):
                continue
            img = cropped.get(label)
            if img is None:
                continue
            px = pl.get("x", 0)
            py = pl.get("y", 0)
            pw = max(1, pl.get("w", IA_W // 2))
            ph = max(1, pl.get("h", PREV_H // 2))

            # 列: コンポーネントは画像エリア(0→IA_W)内に配置するので IA_W で正規化
            c1 = IMG_COL_START + int(px / IA_W * img_cols)
            r1 = IMG_ROW_START + int(py / PREV_H * img_rows)
            c2 = min(IMG_COL_END, c1 + max(1, int(pw / IA_W * img_cols)))
            r2 = min(IMG_ROW_END, r1 + max(1, int(ph / PREV_H * img_rows)))
            placements.append({"img": img, "c1": c1, "r1": r1, "c2": c2, "r2": r2})

    st.session_state["placements"] = placements
    st.session_state["cropped"]    = cropped

# ── Excel 生成 ────────────────────────────────────────────────────────────────
st.divider()
if st.button("📊 不動産案内書 Excel を作成", type="primary", use_container_width=True):
    placements = st.session_state.get("placements", [])

    company_now = load_company_info()
    if ci.get("商号"):
        company_now = ci

    with st.spinner("Excel を生成中..."):
        try:
            xlsx = create_fudosan_excel(
                template_type  = tmpl,
                specs          = edited_specs,
                company_info   = company_now,
                catchphrases   = edited_cps,
                image_placements = placements,
            )
            pname = edited_specs.get("物件名", "物件").replace(" ", "_").replace("/", "_")
            label = "賃貸" if tmpl == "賃貸" else "売買"
            fname = f"{label}案内書_{pname}.xlsx"

            st.download_button(
                label="⬇️ Excel をダウンロード",
                data=xlsx,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )
            st.success("✅ 生成完了！ダウンロードしてください。")
        except Exception as e:
            st.error(f"❌ 生成エラー: {e}")
            st.exception(e)
