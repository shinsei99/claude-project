# -*- coding: utf-8 -*-
"""媒介契約書（一般 / 専任 / 専属専任）を openpyxl で組み立てて xlsx バイト列を返す。

- シート1「媒介契約書」: 表題・締結文・甲乙・本文条項・別表（目的物件の表示）。
  別表は謄本パーサの構造化データを自動転記する（本アプリの中核）。
- シート2「約款」: services.contract_text.YAKKAN の正規約款テキスト。

国土交通省の標準媒介契約約款に基づくフォーマット（templates/original_baikai_template.xls）
の主要部を再構成したもの。
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from services.contract_text import YAKKAN

# ── スタイル定義 ──────────────────────────────────────────────────────────────
FONT = "ＭＳ Ｐ明朝"
FONT_G = "ＭＳ Ｐゴシック"

thin = Side(style="thin", color="000000")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HEAD_FILL = PatternFill("solid", fgColor="E8EEF4")
LABEL_FILL = PatternFill("solid", fgColor="F2F2F2")
FILL_FILL = PatternFill("solid", fgColor="FFFDE7")  # 自動入力箇所の薄い色

WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
WRAP_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

NCOLS = 28  # 別表の桁数に合わせた作業列数


# ── 共通ヘルパ ────────────────────────────────────────────────────────────────
class Sheet:
    """行カーソルを持ち、結合セル・枠・文字を順番に積むラッパ。"""

    def __init__(self, ws):
        self.ws = ws
        self.r = 1

    def _rng(self, r, c1, c2):
        return f"{get_column_letter(c1)}{r}:{get_column_letter(c2)}{r}"

    def cell(self, r, c1, c2, value="", *, font=FONT, size=10, bold=False,
             align=LEFT, fill=None, border=False, color=None, height=None):
        ws = self.ws
        if c2 > c1:
            ws.merge_cells(self._rng(r, c1, c2))
        cell = ws.cell(row=r, column=c1, value=value)
        cell.font = Font(name=font, size=size, bold=bold, color=color)
        cell.alignment = align
        if fill:
            cell.fill = fill
        if border:
            for c in range(c1, c2 + 1):
                bc = ws.cell(row=r, column=c)
                bc.border = BORDER
                if fill and c > c1:
                    bc.fill = fill
        if height:
            ws.row_dimensions[r].height = height
        return cell

    def line(self, c1, c2, value="", **kw):
        """現在行に1行積んでカーソルを進める。"""
        cell = self.cell(self.r, c1, c2, value, **kw)
        self.r += 1
        return cell

    def blank(self, n=1):
        self.r += n


def _yen(n):
    try:
        return "{:,}".format(int(n))
    except (TypeError, ValueError):
        return ""


def _g(d, *keys):
    """ネスト辞書から安全に値を取る。"""
    cur = d
    for k in keys:
        if not isinstance(cur, dict):
            return ""
        cur = cur.get(k, "")
    return cur or ""


# ── 本文条項（種別ごとの差分パラメータ） ──────────────────────────────────────
TYPE_PARAMS = {
    "一般": {
        "title": "一般媒介契約書",
        "yakkan_title": "一般媒介契約約款",
        "report_freq": None,           # 一般は報告頻度の法定義務なし
        "ryutsu_days": None,           # 一般は登録任意
        "max_months": 3,
        "intro": ("この契約は一般媒介契約型式です。依頼者は、目的物件の売買又は交換の媒介又は代理を、"
                  "当社以外の宅地建物取引業者に重ねて依頼することができます。また、自ら発見した相手方と"
                  "売買又は交換の契約を締結することができます。なお、依頼者は、重ねて依頼する宅地建物"
                  "取引業者を明示する義務を負います。"),
    },
    "専任": {
        "title": "専任媒介契約書",
        "yakkan_title": "専任媒介契約約款",
        "report_freq": "２週間に１回以上",
        "ryutsu_days": 7,
        "max_months": 3,
        "intro": ("この契約は専任媒介契約型式です。依頼者は、目的物件の売買又は交換の媒介又は代理を、"
                  "当社以外の宅地建物取引業者に重ねて依頼することができません。依頼者は、自ら発見した"
                  "相手方と売買又は交換の契約を締結することができます。当社は、目的物件を指定流通機構に"
                  "登録します。"),
    },
    "専属専任": {
        "title": "専属専任媒介契約書",
        "yakkan_title": "専属専任媒介契約約款",
        "report_freq": "１週間に１回以上",
        "ryutsu_days": 5,
        "max_months": 3,
        "intro": ("この契約は専属専任媒介契約型式です。依頼者は、目的物件の売買又は交換の媒介又は代理を、"
                  "当社以外の宅地建物取引業者に重ねて依頼することができません。依頼者は、自ら発見した"
                  "相手方と売買又は交換の契約を締結することができません。当社は、目的物件を指定流通機構に"
                  "登録します。"),
    },
}


# ── 別表（目的物件の表示）の描画 ──────────────────────────────────────────────
def _draw_bessou(s: Sheet, data: dict, meta: dict):
    s.line(1, NCOLS, "別表（目的物件の表示）", font=FONT_G, size=11, bold=True,
           align=LEFT, fill=HEAD_FILL, border=True, height=22)

    def kv(label, value, *, fill_value=True, h=26):
        s.cell(s.r, 1, 4, label, align=WRAP_C, fill=LABEL_FILL, border=True, font=FONT_G, size=9.5)
        s.cell(s.r, 5, NCOLS, value, align=WRAP, border=True,
               fill=(FILL_FILL if (fill_value and value) else None), height=h)
        s.r += 1

    # 所有者 / 登記名義人
    kv("所有者　住所", _g(data, "所有者住所"))
    kv("所有者　氏名", _g(data, "所有者氏名"))
    kv("登記名義人　住所", _g(data, "登記名義人住所"))
    kv("登記名義人　氏名", _g(data, "登記名義人氏名"))
    kv("物件所在地", _g(data, "物件所在地"), h=32)

    shubetsu = _g(data, "物件種別")
    if "マンション" in shubetsu:
        _draw_mansion(s, data)
    else:
        _draw_tochi_tatemono(s, data)

    # 媒介価格
    price = meta.get("baikai_price") or 0
    honbody = meta.get("baikai_honbody") or price
    tax = meta.get("baikai_tax") or 0
    s.cell(s.r, 1, 4, "媒介価格", align=WRAP_C, fill=LABEL_FILL, border=True, font=FONT_G, size=9.5)
    s.cell(s.r, 5, 9, "総額", align=CENTER, fill=LABEL_FILL, border=True, size=9)
    s.cell(s.r, 10, 14, (_yen(price) + " 円") if price else "", align=CENTER, border=True,
           fill=(FILL_FILL if price else None))
    s.cell(s.r, 15, 18, "本体価格", align=CENTER, fill=LABEL_FILL, border=True, size=9)
    s.cell(s.r, 19, 22, (_yen(honbody) + " 円") if honbody else "", align=CENTER, border=True)
    s.cell(s.r, 23, 25, "消費税等", align=CENTER, fill=LABEL_FILL, border=True, size=9)
    s.cell(s.r, 26, NCOLS, (_yen(tax) + " 円") if tax else "", align=CENTER, border=True)
    s.r += 1

    kv("備考", meta.get("biko", ""), fill_value=False, h=40)


def _draw_tochi_tatemono(s: Sheet, data: dict):
    t = data.get("土地", {}) or {}
    b = data.get("建物", {}) or {}
    # 土地
    s.cell(s.r, 1, 4, "目的物件\nの表示", align=WRAP_C, fill=LABEL_FILL, border=True, font=FONT_G, size=9)
    s.cell(s.r, 5, 6, "土地", align=CENTER, fill=LABEL_FILL, border=True, size=9)
    s.cell(s.r, 7, 9, "地目", align=CENTER, fill=LABEL_FILL, border=True, size=9)
    s.cell(s.r, 10, 14, _g(t, "地目"), align=CENTER, border=True, fill=(FILL_FILL if _g(t, "地目") else None))
    s.cell(s.r, 15, 16, "権利", align=CENTER, fill=LABEL_FILL, border=True, size=9)
    s.cell(s.r, 17, 21, _g(t, "権利"), align=CENTER, border=True, fill=(FILL_FILL if _g(t, "権利") else None))
    s.cell(s.r, 22, 24, "面積(公簿)", align=CENTER, fill=LABEL_FILL, border=True, size=8.5)
    s.cell(s.r, 25, NCOLS, _g(t, "地積"), align=CENTER, border=True, fill=(FILL_FILL if _g(t, "地積") else None))
    s.r += 1
    # 地番
    s.cell(s.r, 1, 4, "", border=True)
    s.cell(s.r, 5, 9, "地番", align=CENTER, fill=LABEL_FILL, border=True, size=9)
    s.cell(s.r, 10, NCOLS, _g(t, "地番"), align=LEFT, border=True, fill=(FILL_FILL if _g(t, "地番") else None))
    s.r += 1
    # 建物
    s.cell(s.r, 1, 4, "", border=True)
    s.cell(s.r, 5, 6, "建物", align=CENTER, fill=LABEL_FILL, border=True, size=9)
    s.cell(s.r, 7, 9, "種類", align=CENTER, fill=LABEL_FILL, border=True, size=9)
    s.cell(s.r, 10, 16, _g(b, "種類"), align=CENTER, border=True, fill=(FILL_FILL if _g(b, "種類") else None))
    s.cell(s.r, 17, 19, "家屋番号", align=CENTER, fill=LABEL_FILL, border=True, size=8.5)
    s.cell(s.r, 20, NCOLS, _g(b, "家屋番号"), align=CENTER, border=True, fill=(FILL_FILL if _g(b, "家屋番号") else None))
    s.r += 1
    s.cell(s.r, 1, 4, "", border=True)
    s.cell(s.r, 5, 9, "構造", align=CENTER, fill=LABEL_FILL, border=True, size=9)
    s.cell(s.r, 10, NCOLS, _g(b, "構造"), align=LEFT, border=True, fill=(FILL_FILL if _g(b, "構造") else None))
    s.r += 1
    s.cell(s.r, 1, 4, "", border=True)
    s.cell(s.r, 5, 9, "床面積", align=CENTER, fill=LABEL_FILL, border=True, size=9)
    s.cell(s.r, 10, 18, _g(b, "床面積"), align=LEFT, border=True, fill=(FILL_FILL if _g(b, "床面積") else None))
    s.cell(s.r, 19, 21, "新築年月日", align=CENTER, fill=LABEL_FILL, border=True, size=8)
    s.cell(s.r, 22, NCOLS, _g(b, "新築年月日"), align=CENTER, border=True, fill=(FILL_FILL if _g(b, "新築年月日") else None))
    s.r += 1


def _draw_mansion(s: Sheet, data: dict):
    m = data.get("マンション", {}) or {}
    s.cell(s.r, 1, 4, "目的物件\nの表示", align=WRAP_C, fill=LABEL_FILL, border=True, font=FONT_G, size=9)
    s.cell(s.r, 5, 7, "名称", align=CENTER, fill=LABEL_FILL, border=True, size=9)
    s.cell(s.r, 8, NCOLS, _g(m, "名称"), align=LEFT, border=True, fill=(FILL_FILL if _g(m, "名称") else None))
    s.r += 1
    s.cell(s.r, 1, 4, "", border=True)
    s.cell(s.r, 5, 7, "構造", align=CENTER, fill=LABEL_FILL, border=True, size=9)
    s.cell(s.r, 8, 16, _g(m, "構造"), align=LEFT, border=True, fill=(FILL_FILL if _g(m, "構造") else None))
    s.cell(s.r, 17, 20, "階建/階部分", align=CENTER, fill=LABEL_FILL, border=True, size=8)
    val = " / ".join(x for x in [_g(m, "階建"), _g(m, "階部分")] if x)
    s.cell(s.r, 21, NCOLS, val, align=CENTER, border=True, fill=(FILL_FILL if val else None))
    s.r += 1
    s.cell(s.r, 1, 4, "", border=True)
    s.cell(s.r, 5, 7, "専有面積", align=CENTER, fill=LABEL_FILL, border=True, size=8.5)
    s.cell(s.r, 8, 14, _g(m, "専有面積"), align=CENTER, border=True, fill=(FILL_FILL if _g(m, "専有面積") else None))
    s.cell(s.r, 15, 18, "室番号", align=CENTER, fill=LABEL_FILL, border=True, size=8.5)
    s.cell(s.r, 19, NCOLS, _g(m, "室番号"), align=CENTER, border=True, fill=(FILL_FILL if _g(m, "室番号") else None))
    s.r += 1
    s.cell(s.r, 1, 4, "", border=True)
    s.cell(s.r, 5, 7, "新築年月日", align=CENTER, fill=LABEL_FILL, border=True, size=8)
    s.cell(s.r, 8, 16, _g(m, "新築年月日"), align=CENTER, border=True, fill=(FILL_FILL if _g(m, "新築年月日") else None))
    s.cell(s.r, 17, 20, "敷地権割合", align=CENTER, fill=LABEL_FILL, border=True, size=8)
    s.cell(s.r, 21, NCOLS, _g(m, "敷地権割合"), align=CENTER, border=True, fill=(FILL_FILL if _g(m, "敷地権割合") else None))
    s.r += 1
    # 管理費・積立金
    s.cell(s.r, 1, 4, "費用", align=CENTER, fill=LABEL_FILL, border=True, font=FONT_G, size=9)
    s.cell(s.r, 5, 8, "管理費", align=CENTER, fill=LABEL_FILL, border=True, size=8.5)
    s.cell(s.r, 9, 14, "円/月", align=CENTER, border=True)
    s.cell(s.r, 15, 18, "修繕積立金", align=CENTER, fill=LABEL_FILL, border=True, size=8.5)
    s.cell(s.r, 19, NCOLS, "円/月", align=CENTER, border=True)
    s.r += 1


# ── 本文条項の描画 ────────────────────────────────────────────────────────────
def _draw_clauses(s: Sheet, ctype: str, meta: dict):
    p = TYPE_PARAMS[ctype]
    ryutsu_name = meta.get("ryutsu_name", "公益社団法人　不動産流通機構")

    def heading(text):
        s.line(1, NCOLS, text, font=FONT_G, size=10.5, bold=True, fill=HEAD_FILL,
               border=True, align=LEFT, height=20)

    def para(text, indent=2, h=None):
        s.cell(s.r, indent, NCOLS, text, align=WRAP, size=9.5,
               height=h or (15 * (1 + len(text) // 50)))
        s.r += 1

    # 1. 成約に向けての義務
    n = 0
    n += 1
    heading(f"{n}．成約に向けての義務")
    para("乙は、契約の相手方を探索するとともに、契約条件の調整等を行い、契約の成立に向けて積極的に努力します。")
    if p["report_freq"]:
        para(f"乙は、甲に対し、文書又は電子メールにより、{p['report_freq']}の頻度で業務の処理状況を報告します。")
    para("乙は、目的物件の売買又は交換の申込みがあったときは、甲に対し、遅滞なく、その旨を報告します。")
    if p["ryutsu_days"]:
        para(f"乙は、目的物件につき、所在地・規模・形質・媒介価格その他の事項を、{ryutsu_name}に、"
             f"この媒介契約の締結の日の翌日から{p['ryutsu_days']}日以内（乙の休業日を含みません）に登録します。")
    else:
        toroku = "登録します" if meta.get("ryutsu_register", True) else "登録しません"
        para(f"指定流通機構（{ryutsu_name}）への登録の有無：{('有' if meta.get('ryutsu_register', True) else '無')}"
             f"（一般媒介のため任意。{toroku}）。")

    # 2. 媒介に係る業務
    n += 1
    heading(f"{n}．媒介に係る業務")
    para("乙は、媒介価格について意見を述べるときはその根拠を明らかにし、重要事項説明書（宅建業法第35条）及び"
         "契約書面（同第37条）を宅地建物取引士に作成・記名押印・交付させ、登記・決済等の引渡事務の補助を行います。")

    # 3. 建物状況調査のあっせん
    n += 1
    heading(f"{n}．建物状況調査を実施する者のあっせんの有無")
    para(f"（{meta.get('inspection', '無')}）"
         + ("　※既存住宅で「無」とする場合はその理由を記入してください。" if meta.get("inspection", "無") == "無" else ""))

    # 4. 違約金等
    n += 1
    heading(f"{n}．違約金等")
    if ctype == "一般":
        para("甲が通知義務に違反した場合等、乙は約款の定めにより費用の償還を請求することができます。")
    else:
        para("甲が有効期間内に乙以外の宅地建物取引業者に媒介・代理を依頼し売買等を成立させたときは、"
             "乙は約定報酬額に相当する金額（消費税等を除く）を違約金として請求できます。")
        if ctype == "専属専任":
            para("甲が自ら発見した相手方と契約した場合も同様に、乙は約定報酬額に相当する違約金を請求できます。")

    # 5. 有効期間
    n += 1
    heading(f"{n}．有効期間")
    months = meta.get("term_months", "")
    until = meta.get("term_until", "")
    para(f"この媒介契約の有効期間は契約締結後、{months or '　'}ヶ月（{until or '　　年　　月　　日'}まで）とします。"
         f"（{p['max_months']}ヶ月を超えることはできません。）")

    # 6. 約定報酬額
    n += 1
    heading(f"{n}．約定報酬額")
    reward = meta.get("reward") or 0
    reward_tax = meta.get("reward_tax") or 0
    total = (reward + reward_tax) if reward else 0
    para(f"乙の媒介により売買又は交換の契約が成立したときに乙が受領する報酬額は、"
         f"{_yen(reward) + ' 円' if reward else '　　　　　円'}（消費税及び地方消費税抜き報酬額）と"
         f"{_yen(reward_tax) + ' 円' if reward_tax else '　　　円'}（消費税額及び地方消費税額の合計額）を"
         f"合計した額（{_yen(total) + ' 円' if total else '　　　　　円'}）とします。")

    # 7. 約定報酬の受領の時期
    n += 1
    heading(f"{n}．約定報酬の受領の時期")
    para("売買（交換）契約成立時：　　　　／　目的物件の取引完了時：　　　　")

    # 8. 特約事項
    n += 1
    heading(f"{n}．特約事項")
    para("１．甲は乙及び乙の指定する者に、固定資産税台帳の閲覧、評価・公課証明書の取得、ガス・水道等の埋設状況等、"
         "目的物件の重要事項説明等に必要な調査に関する権限を委任します。")
    extra = meta.get("special_terms", "")
    if extra:
        for i, ln in enumerate(extra.splitlines(), start=2):
            if ln.strip():
                para(f"{i}．{ln.strip()}")


# ── 甲乙ブロック ──────────────────────────────────────────────────────────────
def _draw_parties(s: Sheet, meta: dict):
    kou = meta.get("kou", {})
    otsu = meta.get("otsu", {})

    s.cell(s.r, 1, 14, "《依頼者・甲》", font=FONT_G, size=10, bold=True, fill=HEAD_FILL, border=True)
    s.cell(s.r, 15, NCOLS, "《宅地建物取引業者・乙》", font=FONT_G, size=10, bold=True, fill=HEAD_FILL, border=True)
    s.r += 1

    def two(l1, v1, l2, v2, fillL=True):
        s.cell(s.r, 1, 3, l1, align=CENTER, fill=LABEL_FILL, border=True, size=9)
        s.cell(s.r, 4, 14, v1, align=LEFT, border=True, fill=(FILL_FILL if (fillL and v1) else None))
        s.cell(s.r, 15, 18, l2, align=CENTER, fill=LABEL_FILL, border=True, size=9)
        s.cell(s.r, 19, NCOLS, v2, align=LEFT, border=True)
        s.r += 1

    two("住所", kou.get("住所", ""), "商号(名称)", otsu.get("商号", ""))
    two("氏名", kou.get("氏名", ""), "代表者氏名", otsu.get("代表者", ""))
    two("〒/TEL", " ".join(x for x in [kou.get("郵便", ""), kou.get("TEL", "")] if x),
        "所在地", otsu.get("所在地", ""))
    s.cell(s.r, 1, 14, "", border=True)
    s.cell(s.r, 15, 18, "免許番号", align=CENTER, fill=LABEL_FILL, border=True, size=9)
    s.cell(s.r, 19, NCOLS, otsu.get("免許番号", ""), align=LEFT, border=True)
    s.r += 1
    s.cell(s.r, 1, 14, "", border=True)
    s.cell(s.r, 15, 18, "TEL", align=CENTER, fill=LABEL_FILL, border=True, size=9)
    s.cell(s.r, 19, NCOLS, otsu.get("TEL", ""), align=LEFT, border=True)
    s.r += 1


# ── 約款シート ────────────────────────────────────────────────────────────────
def _draw_yakkan(ws, ctype: str):
    s = Sheet(ws)
    paras = YAKKAN.get(ctype, [])
    for p in paras:
        kind = p[0]
        if kind == "title":
            s.line(1, 20, p[1], font=FONT_G, size=13, bold=True, align=CENTER, height=26)
            s.blank()
        elif kind == "header":
            s.line(1, 20, p[1], font=FONT_G, size=10.5, bold=True, height=18)
        elif kind == "article":
            label, body = p[1], p[2]
            s.cell(s.r, 1, 3, label, font=FONT_G, size=10, align=Alignment(horizontal="left", vertical="top"))
            s.cell(s.r, 4, 20, body, align=WRAP, size=10, height=15 * (1 + len(body) // 40))
            s.r += 1
        elif kind == "sub":
            s.cell(s.r, 2, 3, p[1], size=10, align=Alignment(horizontal="left", vertical="top"))
            s.cell(s.r, 4, 20, p[2], align=WRAP, size=10, height=15 * (1 + len(p[2]) // 40))
            s.r += 1
        elif kind == "item":
            s.cell(s.r, 4, 4, p[1], size=10, align=Alignment(horizontal="center", vertical="top"))
            s.cell(s.r, 5, 20, p[2], align=WRAP, size=10, height=15 * (1 + len(p[2]) // 38))
            s.r += 1
        elif kind == "subitem":
            s.cell(s.r, 5, 5, p[1], size=10, align=Alignment(horizontal="center", vertical="top"))
            s.cell(s.r, 6, 20, p[2], align=WRAP, size=10, height=15 * (1 + len(p[2]) // 36))
            s.r += 1
        elif kind == "cont":
            s.cell(s.r, 4, 20, p[1], align=WRAP, size=10, height=15 * (1 + len(p[1]) // 40))
            s.r += 1


# ── 列幅・ページ設定 ──────────────────────────────────────────────────────────
def _setup_page(ws, ncols):
    for c in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 3.4
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.5


# ── エントリポイント ──────────────────────────────────────────────────────────
def build_contract(ctype: str, data: dict, meta: dict) -> bytes:
    """媒介契約書 xlsx を生成してバイト列で返す。

    ctype: "一般" | "専任" | "専属専任"
    data : registry_parser.parse_registry の戻り値（別表用）
    meta : 甲乙・日付・有効期間・媒介価格・報酬・特約 等の入力辞書
    """
    if ctype not in TYPE_PARAMS:
        ctype = "専任"
    p = TYPE_PARAMS[ctype]

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "媒介契約書"
    s = Sheet(ws1)

    # 表題
    s.line(1, NCOLS, p["title"], font=FONT_G, size=16, bold=True, align=CENTER, height=34)
    s.cell(s.r, 1, NCOLS, "この媒介契約は、国土交通省が定めた標準媒介契約約款に基づく契約です。",
           align=CENTER, size=9)
    s.r += 1
    s.cell(s.r, 1, NCOLS, f"依頼の内容：{meta.get('irai_naiyo', '売却')}", align=CENTER, size=9.5, bold=True)
    s.r += 1
    s.blank()

    # 契約型式の説明
    s.cell(s.r, 1, NCOLS, p["intro"], align=WRAP, size=9, height=15 * (1 + len(p["intro"]) // 48),
           border=True, fill=LABEL_FILL)
    s.r += 1
    s.blank()

    # 締結文
    teiketsu = (f"依頼者（以下、甲）は、この契約書及び{p['yakkan_title']}により、別表に表示する不動産"
                "（以下、目的物件）に関する売買（交換）の媒介を下記の宅地建物取引業者（以下、乙）に依頼し、"
                "乙はこれを承諾します。")
    s.cell(s.r, 1, NCOLS, teiketsu, align=WRAP, size=10, height=15 * (1 + len(teiketsu) // 48))
    s.r += 1
    s.cell(s.r, 14, NCOLS, meta.get("date", "令和　　年　　月　　日"), align=Alignment(horizontal="right"), size=10)
    s.r += 1
    s.blank()

    # 甲乙
    _draw_parties(s, meta)
    s.blank()

    # 本文条項
    _draw_clauses(s, ctype, meta)
    s.blank()

    # 別表
    _draw_bessou(s, data, meta)

    _setup_page(ws1, NCOLS)

    # 約款シート
    ws2 = wb.create_sheet("約款")
    _draw_yakkan(ws2, ctype)
    _setup_page(ws2, 20)
    for c in range(1, 21):
        ws2.column_dimensions[get_column_letter(c)].width = 4.0

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
