"""見積書／請求書（Excel）出力サービス。

見積書自動作成ツール（quote-generator）のレイアウトを参考に、原状回復の
「入居者負担額」を賃借人へ提示・請求するための見積書・請求書を生成する。
同じ品目・金額から「見積書」「請求書」の2シートを1冊にまとめて出力できる。
請求書には敷金相殺と振込先欄を追加する。
"""

from __future__ import annotations

import io
import math

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from models.restoration_data import RestorationData


THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

QUOTE = "見積書"
INVOICE = "請求書"

TITLES = {QUOTE: "御　見　積　書", INVOICE: "御　請　求　書"}
GREETINGS = {
    QUOTE: "下記の通り原状回復費用をお見積り申し上げます。",
    INVOICE: "下記の通り原状回復費用をご請求申し上げます。",
}
TOTAL_LABELS = {QUOTE: "御見積金額", INVOICE: "御請求金額"}

# 列構成（A〜F）
COL = {"name": 1, "qty": 2, "unit": 3, "price": 4, "amount": 5, "note": 6}
WIDTHS = {"A": 32, "B": 7, "C": 6, "D": 14, "E": 14, "F": 26}

TAX_RATE = 10  # 表示用（vendor金額は税込前提のため内税として逆算）


def _yen_fmt(cell):
    cell.number_format = "#,##0"


def _build_sheet(wb: Workbook, doc_type: str, data: RestorationData, issuer: dict) -> None:
    ws = wb.create_sheet(doc_type)

    for col, w in WIDTHS.items():
        ws.column_dimensions[col].width = w

    # 入居者負担のある明細のみ請求対象とする
    billable = [it for it in data.items if it.tenant_amount > 0]
    total_incl = sum(it.tenant_amount for it in billable)  # 税込合計
    tax = total_incl - math.floor(total_incl / (1 + TAX_RATE / 100))
    subtotal = total_incl - tax

    # ── タイトル ──
    ws.merge_cells("A1:F2")
    c = ws["A1"]
    c.value = TITLES[doc_type]
    c.font = Font(size=20, bold=True)
    c.alignment = CENTER

    # ── 宛先 ──
    ws.merge_cells("A4:C4")
    ws["A4"] = f"{data.tenant_name} 様"
    ws["A4"].font = Font(size=14)
    ws["A4"].alignment = LEFT
    ws.merge_cells("A5:C5")
    ws["A5"] = f"{data.property_name}　{data.room_number}"
    ws["A5"].alignment = LEFT

    # ── 発行日・発行元 ──
    ws.merge_cells("D4:F4")
    ws["D4"] = f"発行日：{issuer.get('issue_date', '')}"
    ws["D4"].alignment = RIGHT
    ws.merge_cells("D6:F6")
    ws["D6"] = issuer.get("name", "")
    ws["D6"].font = Font(bold=True, size=12)
    ws["D6"].alignment = RIGHT
    ws.merge_cells("D7:F7")
    ws["D7"] = issuer.get("address", "")
    ws["D7"].alignment = Alignment(horizontal="right", wrap_text=True)
    ws.row_dimensions[7].height = 26
    ws.merge_cells("D8:F8")
    tel = issuer.get("tel", "")
    fax = issuer.get("fax", "")
    ws["D8"] = "　".join(x for x in [f"TEL {tel}" if tel else "", f"FAX {fax}" if fax else ""] if x)
    ws["D8"].alignment = RIGHT
    if issuer.get("registration_no"):
        ws.merge_cells("D9:F9")
        ws["D9"] = f"登録番号：{issuer['registration_no']}"
        ws["D9"].alignment = RIGHT
        ws["D9"].font = Font(size=9)

    # ── 挨拶文（誓約書に基づく旨を明記）──
    ws.merge_cells("A8:C9")
    ws["A8"] = (
        GREETINGS[doc_type]
        + "\n※本書は「退去時確認書兼原状回復費用負担誓約書」に基づき作成しています。"
    )
    ws["A8"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[8].height = 30

    # ── 金額サマリーボックス ──
    ws.merge_cells("A11:B11")
    ws["A11"] = TOTAL_LABELS[doc_type]
    ws["A11"].font = Font(bold=True)
    ws["A11"].alignment = CENTER
    ws["A11"].fill = HEADER_FILL
    ws.merge_cells("C11:F11")
    ws["C11"] = total_incl
    ws["C11"].font = Font(size=16, bold=True)
    ws["C11"].alignment = CENTER
    ws["C11"].number_format = '#,##0"円（税込）"'
    for col in range(1, 7):
        ws.cell(row=11, column=col).border = BORDER

    # ── 明細ヘッダー ──
    header_row = 13
    headers = {"name": "工事・部材名", "qty": "数量", "unit": "単位",
               "price": "単価", "amount": "金額", "note": "備考"}
    for key, label in headers.items():
        cell = ws.cell(row=header_row, column=COL[key], value=label)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    # ── 明細行 ──
    row = header_row + 1
    for it in billable:
        ws.cell(row=row, column=COL["name"], value=it.name).alignment = LEFT
        ws.cell(row=row, column=COL["qty"], value=1).alignment = CENTER
        ws.cell(row=row, column=COL["unit"], value="式").alignment = CENTER
        pc = ws.cell(row=row, column=COL["price"], value=it.tenant_amount)
        ac = ws.cell(row=row, column=COL["amount"], value=it.tenant_amount)
        _yen_fmt(pc); _yen_fmt(ac)
        pc.alignment = RIGHT; ac.alignment = RIGHT
        note = f"負担率{it.tenant_rate_pct}%（原状回復）"
        ws.cell(row=row, column=COL["note"], value=note).alignment = LEFT
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = BORDER
        row += 1

    # 空きが少ないときの見栄え用に最低数行ぶん罫線を確保
    min_rows = header_row + 6
    while row <= min_rows:
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = BORDER
        row += 1

    # ── サマリー（小計・消費税・合計） ──
    def _summary_line(r, label, value, bold=False, color=None):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        lc = ws.cell(row=r, column=1, value=label)
        lc.alignment = CENTER
        lc.font = Font(bold=bold, color=color) if color else Font(bold=bold)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=4)
        vc = ws.cell(row=r, column=COL["amount"], value=value)
        _yen_fmt(vc)
        vc.alignment = RIGHT
        vc.font = Font(bold=bold, color=color) if color else Font(bold=bold)
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = BORDER

    _summary_line(row, "小　計（税抜）", subtotal)
    row += 1
    _summary_line(row, f"消費税（{TAX_RATE}%）", tax)
    row += 1
    _summary_line(row, "合　計（税込）", total_incl, bold=True)
    row += 1

    # ── 請求書のみ：敷金相殺・差引請求額・振込先 ──
    if doc_type == INVOICE:
        _summary_line(row, "預り敷金", data.deposit)
        row += 1
        diff = total_incl - data.deposit
        if diff >= 0:
            _summary_line(row, "差引ご請求額", diff, bold=True, color="C00000")
        else:
            _summary_line(row, "敷金ご返還額", -diff, bold=True, color="1F7A1F")
        row += 1

        if issuer.get("bank"):
            row += 1
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1)
            ws.cell(row=row, column=1, value="振込先").font = Font(bold=True)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            ws.cell(row=row, column=2, value=issuer["bank"]).alignment = LEFT

    # ── A4印刷設定 ──
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = "portrait"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    for side in ("left", "right"):
        setattr(ws.page_margins, side, 0.71)
    for side in ("top", "bottom"):
        setattr(ws.page_margins, side, 0.79)


def build(data: RestorationData, issuer: dict, documents: list[str]) -> bytes:
    """指定された書類（見積書/請求書）を1冊のExcelにまとめて返す。

    documents: ["見積書"], ["請求書"], または両方。
    issuer: {name, address, tel, fax, registration_no, bank, issue_date}
    """
    wb = Workbook()
    wb.remove(wb.active)
    for doc in documents:
        if doc in (QUOTE, INVOICE):
            _build_sheet(wb, doc, data, issuer)
    if not wb.sheetnames:
        _build_sheet(wb, QUOTE, data, issuer)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
