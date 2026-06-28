"""査定報告書テンプレート（Excel）を生成する。

3種別ぶんの .xlsx を本フォルダに出力する:
  - satei_mansion.xlsx … 区分マンション用
  - satei_kodate.xlsx  … 土地・戸建用
  - satei_shueki.xlsx  … 収益物件用

データ流し込み先のセル位置は excel_export_service.py のマッピングと一致させる。
テンプレートは「ラベル＋空の入力欄」だけを持ち、値は出力時に書き込む。

  $ python templates/generate_template.py
"""

from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries

HERE = os.path.dirname(__file__)

# スタイル定義
TITLE_FONT = Font(name="游ゴシック", size=18, bold=True)
SECTION_FONT = Font(name="游ゴシック", size=11, bold=True, color="FFFFFF")
LABEL_FONT = Font(name="游ゴシック", size=9, bold=True)
VALUE_FONT = Font(name="游ゴシック", size=10)
BIG_VALUE_FONT = Font(name="游ゴシック", size=16, bold=True, color="C00000")

SECTION_FILL = PatternFill("solid", fgColor="305496")
LABEL_FILL = PatternFill("solid", fgColor="D9E1F2")
VALUE_FILL = PatternFill("solid", fgColor="FFFFFF")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def _new_sheet(title: str, subtitle: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "査定報告書"
    # 列幅（A〜Xを均一の細幅にして方眼ベースのレイアウトにする）
    for col in range(1, 25):  # A..X
        ws.column_dimensions[get_column_letter(col)].width = 4.0
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:X1")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:X2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="游ゴシック", size=9, color="808080")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    return wb, ws


def _section(ws, row: int, text: str, last_col: str = "X"):
    rng = f"A{row}:{last_col}{row}"
    ws.merge_cells(rng)
    c = ws[f"A{row}"]
    c.value = text
    c.font = SECTION_FONT
    c.fill = SECTION_FILL
    c.alignment = LEFT
    ws.row_dimensions[row].height = 20


def _apply_border(ws, rng: str):
    min_c, min_r, max_c, max_r = range_boundaries(rng)
    for r in range(min_r, max_r + 1):
        for c in range(min_c, max_c + 1):
            ws.cell(row=r, column=c).border = BORDER


def _label(ws, rng: str, text: str):
    ws.merge_cells(rng)
    top_left = rng.split(":")[0]
    c = ws[top_left]
    c.value = text
    c.font = LABEL_FONT
    c.fill = LABEL_FILL
    c.alignment = CENTER
    _apply_border(ws, rng)


def _value_box(ws, rng: str, big: bool = False):
    ws.merge_cells(rng)
    top_left = rng.split(":")[0]
    c = ws[top_left]
    c.font = BIG_VALUE_FONT if big else VALUE_FONT
    c.fill = VALUE_FILL
    c.alignment = CENTER if big else LEFT
    _apply_border(ws, rng)
    ws.row_dimensions[int("".join(ch for ch in top_left if ch.isdigit()))].height = 22


def _field(ws, label_rng: str, label: str, value_rng: str, big: bool = False):
    _label(ws, label_rng, label)
    _value_box(ws, value_rng, big=big)


# ---------------------------------------------------------------------------
# 区分マンション
# ---------------------------------------------------------------------------

def build_mansion():
    wb, ws = _new_sheet("不動産査定報告書（区分マンション）", "任売書式-5準拠 ／ 取引事例比較法")

    _section(ws, 4, "■ 対象物件の概要")
    _field(ws, "A5:C5", "マンション名", "D5:S5")
    _field(ws, "T5:U5", "最寄駅", "V5:X5")
    _field(ws, "A6:C6", "物件所在地", "D6:S6")
    _field(ws, "T6:U6", "駅徒歩(分)", "V6:X6")
    _field(ws, "A7:C7", "総戸数", "D7:E7")
    _field(ws, "F7:G7", "階数", "H7:S7")
    _field(ws, "T8:U8", "専有面積(㎡)", "V8:X8")
    _field(ws, "A8:C8", "築年月", "D8:S8")

    _section(ws, 9, "■ 比較事例（近隣中古マンション）")
    # 事例① 行10-13
    _field(ws, "A10:C10", "事例① 名称", "D10:S10")
    _field(ws, "A11:C11", "所在地", "D11:S11")
    _field(ws, "T11:V11", "取引価格(万円)", "W11:X11")
    _field(ws, "T12:V12", "㎡単価(円)", "W12:X12")
    _field(ws, "T13:V13", "取引年月", "W13:X13")
    # 事例② 行16-19
    _field(ws, "A16:C16", "事例② 名称", "D16:S16")
    _field(ws, "A17:C17", "所在地", "D17:S17")
    _field(ws, "T17:V17", "取引価格(万円)", "W17:X17")
    _field(ws, "T18:V18", "㎡単価(円)", "W18:X18")
    _field(ws, "T19:V19", "取引年月", "W19:X19")

    _section(ws, 37, "■ 査定価格")
    _field(ws, "A38:B38", "計算式", "C38:S38")
    _field(ws, "T38:V38", "査定価格", "W38:X38", big=True)

    wb.save(os.path.join(HERE, "satei_mansion.xlsx"))


# ---------------------------------------------------------------------------
# 土地・戸建
# ---------------------------------------------------------------------------

def build_kodate():
    wb, ws = _new_sheet("不動産査定報告書（土地・戸建）", "土地=取引事例比較法 ／ 建物=原価法")

    _section(ws, 4, "■ 対象物件の概要")
    _field(ws, "A5:C5", "所在地", "D5:X5")
    _field(ws, "A6:C6", "土地地積(㎡)", "D6:K6")
    _field(ws, "A7:C7", "延床面積(㎡)", "D7:K7")
    _field(ws, "A8:C8", "構造／築年数", "D8:X8")

    _section(ws, 10, "■ 土地査定根拠（周辺取引事例）")
    _label(ws, "A11:H11", "所在地")
    _label(ws, "I11:P11", "取引価格(万円)")
    _label(ws, "Q11:X11", "㎡単価(円)")
    for i in range(3):
        r = 12 + i
        _value_box(ws, f"A{r}:H{r}")
        _value_box(ws, f"I{r}:P{r}")
        _value_box(ws, f"Q{r}:X{r}")

    _section(ws, 18, "■ 相続税路線価による土地評価（参考）")
    _field(ws, "A19:F19", "採用路線価(円/㎡)", "G19:X19")
    _field(ws, "A20:F20", "算定方法", "G20:X20")
    _field(ws, "A21:F21", "相続税評価額", "G21:L21")
    _field(ws, "M21:Q21", "実勢補正(÷0.8)", "R21:X21")

    _section(ws, 24, "■ 査定価格")
    _label(ws, "A25:V25", "土地評価額（取引事例比較法）")
    _value_box(ws, "W25:X25", big=True)
    _label(ws, "A30:V30", "建物評価額（原価法）")
    _value_box(ws, "W30:X30", big=True)
    _label(ws, "A35:V35", "総額査定価格（土地＋建物）")
    _value_box(ws, "W35:X35", big=True)

    # 算定方法・根拠（土地＝取引事例比較法／建物＝原価法／路線価の内訳）
    _section(ws, 37, "■ 算定方法・根拠")
    ws.merge_cells("A38:X41")
    note = ws["A38"]
    note.font = VALUE_FONT
    note.fill = VALUE_FILL
    note.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    _apply_border(ws, "A38:X41")
    for r in range(38, 42):
        ws.row_dimensions[r].height = 18

    wb.save(os.path.join(HERE, "satei_kodate.xlsx"))


# ---------------------------------------------------------------------------
# 収益物件
# ---------------------------------------------------------------------------

def build_shueki():
    wb, ws = _new_sheet("不動産査定報告書（収益物件・一棟）", "積算価格(コスト法)＋収益価格(収益還元法)")

    _section(ws, 4, "■ 対象物件の概要")
    _field(ws, "A5:C5", "所在地", "D5:X5")
    _field(ws, "A6:C6", "土地地積(㎡)", "D6:K6")
    _field(ws, "A7:C7", "延床面積(㎡)", "D7:K7")
    _field(ws, "A8:C8", "構造／築年数", "D8:X8")
    _field(ws, "A10:C10", "年間想定総収入(万円)", "D10:K10")

    # 左：積算（A〜K）／右：収益（L〜X）を横並びにする
    _section(ws, 13, "■ パターン1：積算価格（コスト法）", last_col="K")
    _field(ws, "A15:G15", "土地積算価格(円)", "H15:K15")
    _field(ws, "A16:G16", "建物積算価格(円)", "H16:K16")
    _field(ws, "A17:G17", "積算合計価格(円)", "H17:K17")

    ws.merge_cells("L13:X13")
    c = ws["L13"]
    c.value = "■ パターン2：収益価格（収益還元法）"
    c.font = SECTION_FONT
    c.fill = SECTION_FILL
    c.alignment = LEFT
    _field(ws, "L15:O15", "年間想定総収入(円)", "P15:X15")
    _field(ws, "L16:O16", "運営経費(年20%・円)", "P16:X16")
    _field(ws, "L17:O17", "期待利回り(%)", "P17:X17")
    _field(ws, "L18:O18", "収益還元価格(円)", "P18:X18")

    _section(ws, 24, "■ 最終査定価格")
    _label(ws, "A25:V25", "最終査定価格（積算と収益の平均）")
    _value_box(ws, "W25:X25", big=True)

    wb.save(os.path.join(HERE, "satei_shueki.xlsx"))


def build_all():
    build_mansion()
    build_kodate()
    build_shueki()
    print("生成しました:", os.listdir(HERE))


if __name__ == "__main__":
    build_all()
