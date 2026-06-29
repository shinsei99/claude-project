# -*- coding: utf-8 -*-
"""3枚セットの査定書Excel（1ブック3シート）を生成する。

  シート1「市場価格分析表」… 査定対象＋取引事例＋周辺売出物件
  シート2「価格査定書」    … 戸建て＝不動産査定書様式 / マンション＝評点様式
  シート3「査定価格の説明書」… 査定根拠の文章

土地戸建て / マンションで様式を分岐。自社情報・ロゴを反映。
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from services.satei_core import (
    TYPE_KODATE, TYPE_MANSION, logo_abspath,
)

# ── スタイル ──────────────────────────────────────────────────────────────────
GOTHIC = "游ゴシック"
MINCHO = "游明朝"
THIN = Side(style="thin", color="000000")
MED = Side(style="medium", color="000000")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BOX_MED = Border(left=MED, right=MED, top=MED, bottom=MED)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)
HEAD_FILL = PatternFill("solid", fgColor="E8EEF7")
TITLE_FILL = PatternFill("solid", fgColor="FFFFFF")


def _c(ws, cell, value="", *, font=MINCHO, size=10.5, bold=False, color="000000",
       align=LEFT, border=None, fill=None):
    cl = ws[cell]
    cl.value = value
    cl.font = Font(name=font, size=size, bold=bold, color=color)
    cl.alignment = align
    if border:
        cl.border = border
    if fill:
        cl.fill = fill
    return cl


def _merge(ws, rng):
    ws.merge_cells(rng)


def _box_range(ws, rng):
    for row in ws[rng]:
        for cell in row:
            cell.border = BOX


def _man(yen) -> str:
    try:
        return f"{round(yen/10000):,}"
    except Exception:
        return ""


def _setup_print(ws, landscape=False):
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.5


def _add_logo(ws, info, anchor="A1", width=120):
    path = logo_abspath(info)
    if not path:
        return
    try:
        img = XLImage(path)
        ratio = img.height / img.width if img.width else 0.5
        img.width = width
        img.height = int(width * ratio)
        ws.add_image(img, anchor)
    except Exception:
        pass


# ── シート1：市場価格分析表 ───────────────────────────────────────────────────
_COLS_KODATE = [
    ("所在地", 22), ("価格\n(万円)", 9), ("土地面積\n(㎡)", 9), ("土地単価\n(円/㎡)", 10),
    ("建物面積\n(㎡)", 9), ("間取り", 8), ("構造", 9), ("築年月", 10),
    ("最寄駅・交通", 18), ("取引年月", 9),
]
_COLS_MANSION = [
    ("所在地・マンション名・号室", 26), ("価格\n(万円)", 9), ("単価\n(円/㎡)", 10),
    ("専有面積\n(㎡)", 9), ("バルコニー\n(㎡)", 9), ("向", 6), ("階／階建", 9),
    ("築年月", 10), ("最寄駅・交通", 16), ("取引年月", 9),
]


def _row_values(ptype, c, with_trade):
    if ptype == TYPE_MANSION:
        vals = [
            (c.get("mansion_name") or c.get("address")), _num(c.get("price_man")),
            _num(c.get("unit_price")), _num(c.get("exclusive_area")),
            _num(c.get("balcony_area")), c.get("direction"), c.get("floor_no"),
            c.get("build_ym"), _access(c),
        ]
    else:
        vals = [
            c.get("address"), _num(c.get("price_man")), _num(c.get("land_area")),
            _num(c.get("unit_price")), _num(c.get("building_area")), c.get("madori"),
            c.get("structure"), c.get("build_ym"), _access(c),
        ]
    vals.append(c.get("trade_ym") if with_trade else "")
    return vals


def _num(v):
    try:
        f = float(v)
        return int(f) if f == int(f) else round(f, 2)
    except Exception:
        return ""


def _access(c):
    s = (c.get("station") or "").strip()
    a = (c.get("access") or "").strip()
    return (s + " " + a).strip()


def _sheet_market(wb, ptype, subject, trades, sales, info, customer, satei_date):
    ws = wb.create_sheet("市場価格分析表")
    cols = _COLS_MANSION if ptype == TYPE_MANSION else _COLS_KODATE
    ncol = len(cols)
    last = get_column_letter(ncol)
    for i, (_, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _add_logo(ws, info, "A1", width=90)
    _c(ws, "A1", "市 場 価 格 分 析 表", font=GOTHIC, size=16, bold=True, align=CENTER)
    _merge(ws, f"A1:{last}1")
    ws.row_dimensions[1].height = 28
    sub = "（マンション用）" if ptype == TYPE_MANSION else "（土地・戸建用）"
    _c(ws, "A2", sub, font=GOTHIC, size=10, align=CENTER)
    _merge(ws, f"A2:{last}2")
    comp = _company_line(info)
    _c(ws, f"A3", comp, font=GOTHIC, size=9, align=RIGHT)
    _merge(ws, f"A3:{last}3")
    _c(ws, "A4", f"査定日：{satei_date}　　お客様：{customer} 様", size=9, align=RIGHT)
    _merge(ws, f"A4:{last}4")

    r = 6

    def section(title, items, with_trade):
        nonlocal r
        _c(ws, f"A{r}", title, font=GOTHIC, size=11, bold=True, align=LEFT)
        _merge(ws, f"A{r}:{last}{r}")
        r += 1
        for i, (h, _) in enumerate(cols, 1):
            _c(ws, f"{get_column_letter(i)}{r}", h, font=GOTHIC, size=9, bold=True,
               align=CENTER, border=BOX, fill=HEAD_FILL)
        ws.row_dimensions[r].height = 26
        r += 1
        rows = items if items else [None]
        for idx, c in enumerate(rows, 1):
            label = ""
            if items and len(items) > 0:
                label = f"{idx}"
            vals = _row_values(ptype, c, with_trade) if c else [""] * ncol
            for i, v in enumerate(vals, 1):
                al = RIGHT if (isinstance(v, (int, float)) and v != "") else LEFT
                _c(ws, f"{get_column_letter(i)}{r}", v, size=9, align=al, border=BOX)
            ws.row_dimensions[r].height = 22
            r += 1
        r += 1

    section("１．査定対象不動産の概要", [subject], with_trade=False)
    section("２．最近の取引事例", trades, with_trade=True)
    section("３．周辺地域の売出物件", sales, with_trade=False)
    _setup_print(ws, landscape=False)
    return ws


def _company_line(info):
    parts = [info.get("company_name", "")]
    if info.get("office"):
        parts.append(info["office"])
    if info.get("staff"):
        parts.append(f"担当：{info['staff']}")
    return "　".join([p for p in parts if p])


# ── シート2：価格査定書 ───────────────────────────────────────────────────────
def _sheet_satei_kodate(wb, subject, trades, plus, minus, units, calc, info, customer, satei_date):
    ws = wb.create_sheet("価格査定書")
    widths = [4, 10, 9, 10, 9, 8, 9, 10, 9, 10, 9, 9]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    last = get_column_letter(len(widths))

    _add_logo(ws, info, "A1", width=80)
    _c(ws, "A1", "不 動 産 査 定 書", font=GOTHIC, size=18, bold=True, color="C00000", align=CENTER)
    _merge(ws, f"A1:{last}1")
    ws.row_dimensions[1].height = 32

    r = 3
    _c(ws, f"A{r}", "お客様氏名", font=GOTHIC, size=9, bold=True, align=CENTER, border=BOX, fill=HEAD_FILL)
    _merge(ws, f"A{r}:B{r}")
    _c(ws, f"C{r}", f"{customer} 様", border=BOX); _merge(ws, f"C{r}:F{r}")
    _c(ws, f"G{r}", "担当者", font=GOTHIC, size=9, bold=True, align=CENTER, border=BOX, fill=HEAD_FILL)
    _c(ws, f"H{r}", info.get("staff", ""), border=BOX); _merge(ws, f"H{r}:I{r}")
    _c(ws, f"J{r}", "査定年月", font=GOTHIC, size=9, bold=True, align=CENTER, border=BOX, fill=HEAD_FILL)
    _c(ws, f"K{r}", satei_date, border=BOX, align=CENTER); _merge(ws, f"K{r}:{last}{r}")
    r += 2

    # 物件情報
    _c(ws, f"A{r}", "物件情報", font=GOTHIC, size=10, bold=True, align=CENTER, border=BOX, fill=HEAD_FILL)
    _merge(ws, f"A{r}:A{r+4}")
    _kv(ws, r, "物件所在地", subject.get("address"), last)
    _kv(ws, r+1, "権利", subject.get("rights"), last)
    _kv(ws, r+2, "最寄駅・交通", _access(subject), last)
    _two(ws, r+3, "土地面積(㎡)", _num(subject.get("land_area")), "建物面積(㎡)", _num(subject.get("building_area")))
    _two(ws, r+4, "構造／間取り", f"{subject.get('structure','')} {subject.get('madori','')}", "築年月", subject.get("build_ym"))
    r += 6

    # 取引事例 ①②③
    _c(ws, f"A{r}", "取引事例（査定価格の算出に採用）", font=GOTHIC, size=10, bold=True, align=LEFT, fill=HEAD_FILL)
    _merge(ws, f"A{r}:{last}{r}")
    r += 1
    head = ["所在地", "取引価格(万円)", "うち土地(万円)", "土地面積(㎡)", "建物面積(㎡)", "構造", "築年月", "最寄駅・交通", "取引年月"]
    cols2 = ["B", "D", "E", "F", "G", "H", "I", "J", "L"]
    _c(ws, f"A{r}", "", border=BOX)
    for h, col in zip(head, cols2):
        _c(ws, f"{col}{r}", h, font=GOTHIC, size=8, bold=True, align=CENTER, border=BOX, fill=HEAD_FILL)
    _merge(ws, f"B{r}:C{r}"); _merge(ws, f"J{r}:K{r}"); _merge(ws, f"L{r}:{last}{r}")
    r += 1
    for idx in range(3):
        c = trades[idx] if idx < len(trades) else None
        vals = [
            (c or {}).get("address", ""), _num((c or {}).get("price_man")),
            _num((c or {}).get("land_price_man")), _num((c or {}).get("land_area")),
            _num((c or {}).get("building_area")), (c or {}).get("structure", ""),
            (c or {}).get("build_ym", ""), _access(c) if c else "", (c or {}).get("trade_ym", ""),
        ]
        _c(ws, f"A{r}", f"{['①','②','③'][idx]}", align=CENTER, border=BOX)
        for v, col in zip(vals, cols2):
            _c(ws, f"{col}{r}", v, size=8, align=LEFT, border=BOX)
        _merge(ws, f"B{r}:C{r}"); _merge(ws, f"J{r}:K{r}"); _merge(ws, f"L{r}:{last}{r}")
        ws.row_dimensions[r].height = 20
        r += 1
    r += 1

    # 加点・減点ポイント
    r = _points_block(ws, r, plus, minus, last)

    # 査定結果
    _c(ws, f"A{r}", "査定結果", font=GOTHIC, size=10, bold=True, align=CENTER, border=BOX, fill=HEAD_FILL)
    _merge(ws, f"A{r}:A{r+3}")
    lu, la, lp = calc["land_unit"], calc["land_area"], calc["land_point"]
    bu, ba, bp = calc["building_unit"], calc["building_area"], calc["building_point"]
    _c(ws, f"B{r}", f"(土地) 土地事例単価 {int(lu):,}円/㎡ × {la:g}㎡ × (100{lp:+d})/100", size=9, border=BOX)
    _merge(ws, f"B{r}:I{r}")
    _c(ws, f"J{r}", "土地価格(A)", font=GOTHIC, size=9, bold=True, align=CENTER, border=BOX)
    _c(ws, f"K{r}", f"{calc['land_value']:,} 円", align=RIGHT, border=BOX); _merge(ws, f"K{r}:{last}{r}")
    r += 1
    _c(ws, f"B{r}", f"(建物) 再調達単価 {int(bu):,}円/㎡ × {ba:g}㎡ × (100{bp:+d})/100", size=9, border=BOX)
    _merge(ws, f"B{r}:I{r}")
    _c(ws, f"J{r}", "建物価格(B)", font=GOTHIC, size=9, bold=True, align=CENTER, border=BOX)
    _c(ws, f"K{r}", f"{calc['building_value']:,} 円", align=RIGHT, border=BOX); _merge(ws, f"K{r}:{last}{r}")
    r += 1
    rr = calc.get("ryutsu", 100)
    _c(ws, f"B{r}", f"（土地A ＋ 建物B）= {calc.get('base', calc['total']):,}円　× 流通性比率 {rr:g}%",
       font=GOTHIC, size=10, align=CENTER, border=BOX)
    _merge(ws, f"B{r}:I{r+1}")
    _c(ws, f"J{r}", "査定価格", font=GOTHIC, size=11, bold=True, align=CENTER, border=BOX_MED, fill=HEAD_FILL)
    _merge(ws, f"J{r}:J{r+1}")
    _c(ws, f"K{r}", f"{calc['total']:,} 円", font=GOTHIC, size=13, bold=True, color="C00000", align=CENTER, border=BOX_MED)
    _merge(ws, f"K{r}:{last}{r+1}")
    ws.row_dimensions[r].height = 22; ws.row_dimensions[r+1].height = 22
    _setup_print(ws, landscape=False)
    return ws


def _sheet_satei_mansion(wb, subject, trades, plus, minus, units, calc, info, customer, satei_date):
    ws = wb.create_sheet("価格査定書")
    widths = [4, 14, 12, 12, 12, 12, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    last = get_column_letter(len(widths))

    _add_logo(ws, info, "A1", width=80)
    _c(ws, "A1", "価 格 査 定 書", font=GOTHIC, size=16, bold=True, align=CENTER)
    _merge(ws, f"A1:{last}1"); ws.row_dimensions[1].height = 28
    _c(ws, "A2", "（中古マンション用）", font=GOTHIC, size=10, align=CENTER); _merge(ws, f"A2:{last}2")

    r = 4
    _c(ws, f"A{r}", "お客さま氏名", font=GOTHIC, size=9, bold=True, align=CENTER, border=BOX, fill=HEAD_FILL)
    _merge(ws, f"A{r}:B{r}")
    _c(ws, f"C{r}", f"{customer} 様", border=BOX); _merge(ws, f"C{r}:E{r}")
    _c(ws, f"F{r}", "査定日", font=GOTHIC, size=9, bold=True, align=CENTER, border=BOX, fill=HEAD_FILL)
    _c(ws, f"G{r}", satei_date, align=CENTER, border=BOX); _merge(ws, f"G{r}:{last}{r}")
    r += 2

    _c(ws, f"A{r}", "査定対象・事例マンション", font=GOTHIC, size=10, bold=True, align=LEFT)
    _merge(ws, f"A{r}:{last}{r}"); r += 1
    head = ["区分", "マンション名・号室", "所在地", "専有面積(㎡)", "単価(円/㎡)", "築年月", "向", "階／階建"]
    for i, h in enumerate(head, 1):
        _c(ws, f"{get_column_letter(i)}{r}", h, font=GOTHIC, size=9, bold=True, align=CENTER, border=BOX, fill=HEAD_FILL)
    ws.row_dimensions[r].height = 24; r += 1
    rows = [("査定", subject)] + [(f"事例{['①','②','③'][i]}", t) for i, t in enumerate(trades[:3])]
    for label, c in rows:
        vals = [label, c.get("mansion_name") or c.get("address"), c.get("address"),
                _num(c.get("exclusive_area")), _num(c.get("unit_price")), c.get("build_ym"),
                c.get("direction"), c.get("floor_no")]
        for i, v in enumerate(vals, 1):
            _c(ws, f"{get_column_letter(i)}{r}", v, size=9, align=LEFT, border=BOX)
        ws.row_dimensions[r].height = 20; r += 1
    r += 1

    r = _points_block(ws, r, plus, minus, last)

    _c(ws, f"A{r}", "査定結果", font=GOTHIC, size=10, bold=True, align=CENTER, border=BOX, fill=HEAD_FILL)
    _merge(ws, f"A{r}:A{r+1}")
    cu, ar, pt = calc["case_unit"], calc["exclusive_area"], calc["point"]
    rr = calc.get("ryutsu", 100)
    _c(ws, f"B{r}", f"事例単価 {int(cu):,}円/㎡ × (100{pt:+d})/100 × 専有面積 {ar:g}㎡\n"
                    f"= {calc.get('base', calc['total']):,}円 × 流通性比率 {rr:g}%", size=10, border=BOX)
    _merge(ws, f"B{r}:E{r+1}")
    _c(ws, f"F{r}", "査定価格", font=GOTHIC, size=11, bold=True, align=CENTER, border=BOX_MED, fill=HEAD_FILL)
    _merge(ws, f"F{r}:F{r+1}")
    _c(ws, f"G{r}", f"{calc['total']:,} 円", font=GOTHIC, size=13, bold=True, color="C00000", align=CENTER, border=BOX_MED)
    _merge(ws, f"G{r}:{last}{r+1}")
    ws.row_dimensions[r].height = 22; ws.row_dimensions[r+1].height = 22
    _setup_print(ws, landscape=False)
    return ws


def _points_block(ws, r, plus, minus, last):
    _c(ws, f"A{r}", "加点ポイント", font=GOTHIC, size=9, bold=True, align=CENTER, border=BOX, fill=HEAD_FILL)
    _merge(ws, f"A{r}:E{r}")
    _c(ws, f"F{r}", "減点ポイント", font=GOTHIC, size=9, bold=True, align=CENTER, border=BOX, fill=HEAD_FILL)
    _merge(ws, f"F{r}:{last}{r}")
    r += 1
    for col0, col_factor, col_kubun, col_pt in [("A", "B", "D", "E")]:
        _c(ws, f"{col0}{r}", "", border=BOX)
        _c(ws, f"{col_factor}{r}", "要因", font=GOTHIC, size=8, bold=True, align=CENTER, border=BOX); _merge(ws, f"B{r}:C{r}")
        _c(ws, f"{col_kubun}{r}", "区分", font=GOTHIC, size=8, bold=True, align=CENTER, border=BOX)
        _c(ws, f"{col_pt}{r}", "点", font=GOTHIC, size=8, bold=True, align=CENTER, border=BOX)
    _c(ws, f"F{r}", "", border=BOX)
    _c(ws, f"G{r}", "要因", font=GOTHIC, size=8, bold=True, align=CENTER, border=BOX); _merge(ws, f"G{r}:{last}{r}")
    # right side kubun/point compacted into one note column for space
    r += 1
    for i in range(5):
        mark = ["①", "②", "③", "④", "⑤"][i]
        p = plus[i] if i < len(plus) else None
        m = minus[i] if i < len(minus) else None
        _c(ws, f"A{r}", mark, align=CENTER, border=BOX)
        _c(ws, f"B{r}", (p or {}).get("factor", ""), size=8, border=BOX); _merge(ws, f"B{r}:C{r}")
        _c(ws, f"D{r}", (p or {}).get("kubun", "") if p and p.get("factor") else "", size=8, align=CENTER, border=BOX)
        _c(ws, f"E{r}", (p or {}).get("point", "") if p and p.get("factor") else "", size=8, align=CENTER, border=BOX)
        _c(ws, f"F{r}", mark, align=CENTER, border=BOX)
        mtxt = ""
        if m and m.get("factor"):
            mtxt = f"{m.get('factor','')}（{m.get('kubun','')} -{m.get('point','')}）"
        _c(ws, f"G{r}", mtxt, size=8, border=BOX); _merge(ws, f"G{r}:{last}{r}")
        r += 1
    return r + 1


def _kv(ws, r, key, val, last):
    _c(ws, f"B{r}", key, font=GOTHIC, size=9, bold=True, align=CENTER, border=BOX, fill=HEAD_FILL)
    _merge(ws, f"B{r}:C{r}")
    _c(ws, f"D{r}", val, size=9, border=BOX); _merge(ws, f"D{r}:{last}{r}")


def _two(ws, r, k1, v1, k2, v2):
    _c(ws, f"B{r}", k1, font=GOTHIC, size=9, bold=True, align=CENTER, border=BOX, fill=HEAD_FILL); _merge(ws, f"B{r}:C{r}")
    _c(ws, f"D{r}", v1, size=9, align=RIGHT, border=BOX); _merge(ws, f"D{r}:F{r}")
    _c(ws, f"G{r}", k2, font=GOTHIC, size=9, bold=True, align=CENTER, border=BOX, fill=HEAD_FILL); _merge(ws, f"G{r}:H{r}")
    _c(ws, f"I{r}", v2, size=9, align=RIGHT, border=BOX); _merge(ws, f"I{r}:L{r}")


# ── シート3：査定価格の説明書 ─────────────────────────────────────────────────
def _sheet_explanation(wb, subject, calc, info, customer, satei_date, expiry, explanation):
    ws = wb.create_sheet("査定価格の説明書")
    for i, w in enumerate([4, 14, 14, 14, 14, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    last = "H"
    _add_logo(ws, info, "A1", width=80)
    r = 2
    _c(ws, f"A{r}", satei_date, size=10, align=RIGHT); _merge(ws, f"A{r}:{last}{r}"); r += 1
    _c(ws, f"A{r}", f"{customer}　様", font=GOTHIC, size=12, bold=True, align=LEFT); _merge(ws, f"A{r}:E{r}"); r += 1
    _c(ws, f"E{r}", _company_line(info), font=GOTHIC, size=10, align=RIGHT); _merge(ws, f"D{r}:{last}{r}"); r += 2
    _c(ws, f"A{r}", "査定価格についての説明書", font=GOTHIC, size=15, bold=True, align=CENTER); _merge(ws, f"A{r}:{last}{r}")
    ws.row_dimensions[r].height = 30; r += 2

    intro = ("この度は、当社にお問い合わせを賜りありがとうございます。ご売却に向けての指針として、"
             "現在の査定価格を算出いたしました。ご参考にして頂ければ幸いです。\n"
             "尚、この査定価格は通常３ヶ月以内で売却できると思われる価格です。現地調査の上、取引事例・"
             "販売中の物件等の資料を検討し、合理的かつ適正な意見価格としてご報告申し上げるものです。")
    _c(ws, f"A{r}", intro, size=10.5, align=LEFT); _merge(ws, f"A{r}:{last}{r+3}")
    ws.row_dimensions[r].height = 70; r += 5

    _c(ws, f"A{r}", "査定物件", font=GOTHIC, size=11, bold=True, border=BOX, fill=HEAD_FILL); _merge(ws, f"A{r}:B{r}")
    _c(ws, f"C{r}", subject.get("address", ""), border=BOX); _merge(ws, f"C{r}:{last}{r}"); r += 1
    area = subject.get("exclusive_area") or subject.get("land_area") or 0
    area_label = "専有面積" if calc.get("type") == TYPE_MANSION else "土地面積"
    tsubo = f"（約{round(area/3.30578,2):g}坪）" if area else ""
    _c(ws, f"A{r}", area_label, font=GOTHIC, size=11, bold=True, border=BOX, fill=HEAD_FILL); _merge(ws, f"A{r}:B{r}")
    _c(ws, f"C{r}", f"{area:g}㎡ {tsubo}", border=BOX); _merge(ws, f"C{r}:{last}{r}"); r += 2

    _c(ws, f"A{r}", "査定価格", font=GOTHIC, size=13, bold=True, align=CENTER, border=BOX_MED, fill=HEAD_FILL); _merge(ws, f"A{r}:B{r}")
    _c(ws, f"C{r}", f"{calc['total']:,} 円", font=GOTHIC, size=15, bold=True, color="C00000", align=CENTER, border=BOX_MED); _merge(ws, f"C{r}:E{r}")
    _c(ws, f"F{r}", f"有効期限：{expiry}", size=10, align=CENTER, border=BOX_MED); _merge(ws, f"F{r}:{last}{r}")
    ws.row_dimensions[r].height = 30; r += 2

    _c(ws, f"A{r}", "査定の根拠", font=GOTHIC, size=12, bold=True, align=LEFT); _merge(ws, f"A{r}:{last}{r}"); r += 1
    _c(ws, f"A{r}", explanation or "（査定の根拠を入力／AI生成してください）", size=10.5, align=LEFT)
    _merge(ws, f"A{r}:{last}{r+6}")
    ws.row_dimensions[r].height = 150; r += 8

    _c(ws, f"A{r}", "以上", size=10.5, align=RIGHT); _merge(ws, f"A{r}:{last}{r}")
    _setup_print(ws, landscape=False)
    return ws


# ── エントリポイント ──────────────────────────────────────────────────────────
def build_report(
    *, property_type, subject, trades, sales, plus, minus, units, calc,
    company, customer, satei_date, expiry, explanation,
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    _sheet_market(wb, property_type, subject, trades, sales, company, customer, satei_date)
    if property_type == TYPE_MANSION:
        _sheet_satei_mansion(wb, subject, trades, plus, minus, units, calc, company, customer, satei_date)
    else:
        _sheet_satei_kodate(wb, subject, trades, plus, minus, units, calc, company, customer, satei_date)
    _sheet_explanation(wb, subject, calc, company, customer, satei_date, expiry, explanation)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
