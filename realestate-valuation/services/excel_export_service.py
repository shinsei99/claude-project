"""査定報告書（Excel）出力サービス。

物件種別ごとに `templates/satei_*.xlsx` を読み込み、openpyxl で
仕様どおりの指定セルにデータを流し込む。テンプレートのレイアウト・
罫線・結合は維持したまま、値だけを書き込む。
"""

from __future__ import annotations

import io
import os

from openpyxl import load_workbook

from models.valuation_data import (
    ValuationPipelineData,
    TYPE_MANSION,
    TYPE_KODATE,
    TYPE_SHUEKI,
)

_TEMPLATE_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "templates")

_TEMPLATE_FILE = {
    TYPE_MANSION: "satei_mansion.xlsx",
    TYPE_KODATE: "satei_kodate.xlsx",
    TYPE_SHUEKI: "satei_shueki.xlsx",
}


def _man_str(yen: int) -> str:
    """円 → 「○○万円」表記。0なら空文字。"""
    if not yen:
        return ""
    return f"{round(yen / 10000):,}万円"


def _put(ws, cell: str, value) -> None:
    """結合セルの左上に値を書き込む（空値はスキップ）。"""
    if value in (None, "", 0, 0.0):
        return
    ws[cell] = value


def _struct_age(reg) -> str:
    age = ""
    if reg.build_year:
        from datetime import date

        age = f"（築{max(0, date.today().year - reg.build_year)}年）"
    return f"{reg.structure}{age}".strip()


def _fill_mansion(ws, data: ValuationPipelineData) -> None:
    reg, v, m = data.registry, data.valuation, data.market
    _put(ws, "D5", reg.mansion_name)
    _put(ws, "V5", reg.nearest_station)
    _put(ws, "D6", data.address or reg.location)
    _put(ws, "V6", reg.station_minutes)
    _put(ws, "D7", reg.total_units)
    if reg.floor_no or reg.total_floors:
        ws["H7"] = f"{reg.floor_no}階／{reg.total_floors}階建"
    _put(ws, "D8", reg.build_ym)
    _put(ws, "V8", reg.exclusive_area or reg.floor_area)

    # 比較事例（最大2件を仕様セルへ）
    comp_cells = [
        {"name": "D10", "addr": "D11", "price": "W11", "unit": "W12", "period": "W13"},
        {"name": "D16", "addr": "D17", "price": "W17", "unit": "W18", "period": "W19"},
    ]
    for c, cell in zip(m.comparables, comp_cells):
        _put(ws, cell["name"], c.name)
        _put(ws, cell["addr"], c.address)
        _put(ws, cell["price"], c.trade_price_man)
        _put(ws, cell["unit"], c.unit_price)
        _put(ws, cell["period"], c.trade_period)

    _put(ws, "C38", v.basis)
    _put(ws, "W38", _man_str(v.final_price))


def _fill_kodate(ws, data: ValuationPipelineData) -> None:
    reg, v, m = data.registry, data.valuation, data.market
    _put(ws, "D5", data.address or reg.location)
    _put(ws, "D6", reg.land_area)
    _put(ws, "D7", reg.floor_area)
    _put(ws, "D8", _struct_age(reg))

    # 土地取引事例（最大3件）: 住所 / 取引価格(万円) / ㎡単価
    for i, c in enumerate(m.comparables[:3]):
        r = 12 + i
        _put(ws, f"A{r}", c.address or c.name)
        _put(ws, f"I{r}", c.trade_price_man)
        _put(ws, f"Q{r}", c.unit_price)

    _put(ws, "W25", _man_str(v.land_price))
    _put(ws, "W30", _man_str(v.building_price))
    _put(ws, "W35", _man_str(v.final_price))

    # 相続税路線価による土地評価（参考）
    _put(ws, "G19", v.rosenka_unit_price)
    _put(ws, "G20", v.rosenka_detail)
    _put(ws, "G21", _man_str(v.rosenka_souzoku))
    _put(ws, "R21", _man_str(v.rosenka_jissei))

    # 算定方法・根拠
    _put(ws, "A38", v.basis)


def _fill_shueki(ws, data: ValuationPipelineData) -> None:
    reg, v = data.registry, data.valuation
    _put(ws, "D5", data.address or reg.location)
    _put(ws, "D6", reg.land_area)
    _put(ws, "D7", reg.floor_area)
    _put(ws, "D8", _struct_age(reg))
    _put(ws, "D10", data.rentroll.annual_income_man)

    # パターン1：積算（円）
    _put(ws, "H15", v.cost_land)
    _put(ws, "H16", v.cost_building)
    _put(ws, "H17", v.cost_total)

    # パターン2：収益（円）。運営経費はマイナス表示。
    _put(ws, "P15", v.income_gross)
    if v.income_expense:
        ws["P16"] = -v.income_expense
    _put(ws, "P17", v.cap_rate)
    _put(ws, "P18", v.income_price)

    _put(ws, "W25", _man_str(v.final_price))


_FILLERS = {
    TYPE_MANSION: _fill_mansion,
    TYPE_KODATE: _fill_kodate,
    TYPE_SHUEKI: _fill_shueki,
}


def build(data: ValuationPipelineData) -> bytes:
    """査定報告書Excelを生成してバイト列で返す。"""
    fname = _TEMPLATE_FILE.get(data.property_type)
    if not fname:
        raise ValueError(f"未知の物件種別: {data.property_type}")
    path = os.path.join(_TEMPLATE_DIR, fname)
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"テンプレートが見つかりません: {path}\n"
            "`python templates/generate_template.py` を実行して生成してください。"
        )

    wb = load_workbook(path)
    ws = wb.active
    _FILLERS[data.property_type](ws, data)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
